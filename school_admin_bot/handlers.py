from aiogram import Bot, F, Router
from aiogram.types import CallbackQuery, InlineKeyboardButton, InlineKeyboardMarkup, Message
from aiogram.filters import Command, CommandStart
from aiogram.fsm.context import FSMContext
import asyncio
import logging
import os
import re
import time
from datetime import date, datetime, timedelta, timezone
from pathlib import Path
from uuid import uuid4
from zoneinfo import ZoneInfo

from config import MAX_BOT_TOKEN, SCHOOL_BOT_PAYMENTS_CHAT_ID, SCHOOL_BOT_TOKEN, SCHOOL_BOT_USERNAME, SUPERADMINS

from _utils import (
    BOT_DIR,
    MSK_TZ,
    PROJECT_ROOT,
    TEACHER_UPLOADS_DIR,
    build_payment_prompt_keyboard,
    get_admin_reply_menu,
    get_home_menu_by_user_id,
    get_role_by_user_id,
    is_admin_role,
    is_teacher_role,
    msk_now_naive,
    notify_student_about_attendance,
    notify_teacher_about_attendance,
    resolve_local_path,
    role_title,
    save_teacher_photo,
    send_student_notification,
    update_flow_message,
)


from keyboards import (
    get_superadmin_menu,
    get_superadmin_users_menu,
    get_superadmin_school_menu,
    get_superadmin_reports_menu,
    get_admin_menu,
    get_admin_management_menu,
    get_admin_education_menu,
    get_admin_finance_menu,
    get_admin_reports_menu,
    get_teacher_menu,
    get_student_menu,
    get_tariff_keyboard,
    get_attendance_direction_keyboard,
    get_attendance_mark_keyboard,
    get_teacher_attendance_students_keyboard,
    get_balance_direction_keyboard,
    get_balance_add_keyboard,
    get_teacher_bind_keyboard,
    get_role_change_keyboard,
    get_main_menu_shortcut_keyboard,
    get_user_selection_keyboard,
    get_teacher_selection_keyboard,
    get_student_disambiguation_keyboard,
    get_subject_selection_keyboard,
    get_assign_subject_rename_keyboard,
    get_teacher_subject_picker_keyboard,
    get_edit_teacher_subject_picker_keyboard,
    get_publication_audience_keyboard,
    get_publication_schedule_keyboard,
    get_lessons_report_period_keyboard,
    get_lessons_report_teacher_filter_keyboard,
    get_promo_discount_type_kb,
)
from states import AdminStates
from shared.database import (
    add_student,
    get_all_students,
    add_teacher_if_not_exists,
    add_student_lesson,
    find_students_by_name_with_username,
    get_student_directions,
    get_student_lesson_by_id,
    mark_attendance,
    has_recent_attendance,
    add_lessons_to_balance,
    get_balance_history_by_student,
    add_user,
    get_user_by_telegram_id,
    get_users_by_role,
    search_users_by_name_or_username,
    get_user_by_id,
    get_student_by_id,
    get_student_by_id_with_username,
    get_student_by_telegram_id,
    bind_teacher_telegram_id,
    log_admin_action,
    get_recent_admin_actions,
    format_admin_action_log,
    get_teacher_weekly_lessons_report,
    get_weekly_lessons_report_for_teacher_telegram,
    get_teacher_lessons_report,
    get_teachers_with_lessons,
    format_teacher_weekly_report,
    get_recent_payment_history_by_telegram_user,
    build_daily_debt_report,
    get_students_by_teacher_telegram_id,
    get_teacher_by_telegram_id,
    get_teacher_by_id,
    search_teacher_profiles,
    list_teacher_profiles,
    get_teacher_profile_by_id,
    get_teacher_catalog_subjects,
    update_teacher_profile_fields,
    set_teacher_telegram_id,
    update_user_role,
    set_user_active,
    delete_admin_by_telegram_id,
    delete_teacher_by_telegram_id,
    delete_student_by_telegram_id,
    add_or_update_teacher_profile,
    get_active_admin_contacts,
    get_known_telegram_user_id_by_username,
    create_onboarding_invite,
    normalize_telegram_username,
    get_latest_pending_invite_by_role_and_username,
    mark_onboarding_invite_used,
    upsert_known_telegram_user,
    create_publication_post,
    create_review_card,
    get_active_review_cards,
    deactivate_review_card,
    update_review_card_media,
    get_admin_dashboard_metrics,
    get_current_debtors_summary,
    get_debtor_student_details,
    get_referral_by_invitee_telegram_id,
    link_invitee_student,
    get_student_max_id,
    create_promo_code,
    assign_promo_to_student,
    list_promo_codes,
    get_promo_code_by_id,
    deactivate_promo_code,
    activate_promo_code,
    delete_promo_code,
)
from shared.sheets import get_sheets_client
from shared.database import (
    sheets_outbox_add,
    sheets_outbox_get_dead,
    sheets_outbox_delete_dead,
    get_weekly_payouts,
    get_all_student_balances,
    get_attendance_stats,
    get_revenue_by_period,
    get_topups_history,
)

router = Router()
router.message.filter(F.chat.type == "private")
router.callback_query.filter(F.message.chat.type == "private")
logger = logging.getLogger(__name__)

_summary_sheets_lock = asyncio.Lock()


async def _update_summary_sheets_bg() -> None:
    """Background: refresh Выплаты, Балансы, Статистика in Google Sheets.

    Uses a lock so parallel attendance marks don't write the sheets simultaneously.
    If an update is already running, this call is skipped (не накапливаются в очередь).
    """
    if _summary_sheets_lock.locked():
        logger.debug("Sheets: summary update already running, skipping")
        return
    async with _summary_sheets_lock:
        client = get_sheets_client()
        if not client.is_configured():
            return
        for fetch_fn, update_fn, label in [
            (get_weekly_payouts,       client.update_payouts_sheet,  "Выплаты"),
            (get_all_student_balances, client.update_balances_sheet, "Балансы"),
            (get_attendance_stats,     client.update_stats_sheet,    "Статистика"),
            (get_revenue_by_period,    client.update_revenue_sheet,  "Выручка"),
            (get_topups_history,       client.update_topups_sheet,   "Пополнения"),
        ]:
            try:
                data = await asyncio.to_thread(fetch_fn)
                await asyncio.to_thread(update_fn, data)
                logger.info("Sheets: %s refreshed", label)
            except Exception as exc:
                logger.warning("Sheets: %s refresh failed: %s", label, exc)


async def _sync_attendance_to_sheets(
    attendance_id: int,
    lesson_datetime: str,
    teacher_name: str,
    student_name: str,
    subject_name: str,
    tariff_type: str,
    status: str,
    balance_before: int,
    balance_after: int,
    marked_by_name: str,
) -> None:
    """Fire-and-forget: push one attendance row to Google Sheets.

    On failure the row is saved to sheets_outbox and retried by
    _flush_sheets_outbox. The DB write always succeeds regardless.
    """
    client = get_sheets_client()
    if not client.is_configured():
        return
    row = {
        "attendance_id": attendance_id,
        "lesson_datetime": lesson_datetime,
        "teacher_name": teacher_name,
        "student_name": student_name,
        "subject_name": subject_name,
        "tariff_type": tariff_type,
        "status": status,
        "balance_before": balance_before,
        "balance_after": balance_after,
        "marked_by_name": marked_by_name,
    }
    try:
        ok = await asyncio.to_thread(client.append_attendance, row)
        if not ok:
            raise RuntimeError("append_attendance returned False")
    except Exception as exc:
        logger.warning("Sheets sync failed, queuing for retry (attendance_id=%s): %s", attendance_id, exc)
        await asyncio.to_thread(sheets_outbox_add, attendance_id, row)



def build_payment_prompt_keyboard_clean() -> InlineKeyboardMarkup | None:
    if not SCHOOL_BOT_USERNAME:
        return None
    return InlineKeyboardMarkup(
        inline_keyboard=[
            [
                InlineKeyboardButton(
                    text="💸 Погасить долг",
                    url=f"https://t.me/{SCHOOL_BOT_USERNAME}?start=pay_debt",
                )
            ]
        ]
    )


async def _send_max_notification(max_user_id: int, text: str, max_kb=None) -> None:
    if not MAX_BOT_TOKEN:
        return
    try:
        from shared.max_api import MaxApiClient
        api = MaxApiClient(MAX_BOT_TOKEN)
        await api.send_message(max_user_id, text, max_kb)
    except Exception as exc:
        logger.warning("Failed to send MAX notification to user %s: %s", max_user_id, exc)


async def notify_student_about_attendance_clean(
    callback: CallbackQuery,
    *,
    student_telegram_id: int | None,
    student_max_id: int | None = None,
    student_name: str,
    subject_name: str,
    teacher_name: str,
    tariff_type: str,
    status: str,
    lesson_balance_before: int,
    lesson_balance_after: int,
) -> None:
    if not student_telegram_id and not student_max_id:
        return

    if status != "present":
        text = (
            "Здравствуйте!\n\n"
            "По Вашему направлению обновлена отметка посещаемости.\n\n"
            f"Ученик: {student_name}\n"
            f"Предмет: {subject_name}\n"
            f"Преподаватель: {teacher_name}\n"
            "Статус занятия: не был."
        )
        if student_telegram_id:
            await send_student_notification(callback, student_telegram_id, text)
        if student_max_id:
            from shared.max_api import btn as _mbtn, keyboard as _mkb
            max_kb = _mkb([_mbtn("👤 Личный кабинет", "menu_cabinet")])
            await _send_max_notification(student_max_id, text, max_kb)
        return

    lines = [
        "Здравствуйте!",
        "",
        "Занятие отмечено как проведённое.",
        "",
        f"Ученик: {student_name}",
        f"Предмет: {subject_name}",
        f"Преподаватель: {teacher_name}",
        "Списано занятий: 1",
        f"Баланс был: {lesson_balance_before}",
        f"Баланс стал: {lesson_balance_after}",
    ]

    reply_markup = None
    max_kb = None
    if lesson_balance_after < 0:
        lines.extend(
            [
                "",
                "❗❗❗🔴 ВНИМАНИЕ! У ВАС ЗАДОЛЖЕННОСТЬ! 🔴❗❗❗",
                f"Размер задолженности: {abs(lesson_balance_after)} занят.",
                "❗❗❗ Пожалуйста, внесите оплату. ❗❗❗",
            ]
        )
        reply_markup = build_payment_prompt_keyboard_clean()
        from shared.max_api import btn as _mbtn, keyboard as _mkb
        max_kb = _mkb(
            [_mbtn("👤 Личный кабинет", "menu_cabinet")],
            [_mbtn("💸 Погасить долг", "menu_paid")],
        )
    elif lesson_balance_after == 0:
        lines.extend(["", "На балансе больше не осталось оплаченных занятий."])
        from shared.max_api import btn as _mbtn, keyboard as _mkb
        max_kb = _mkb([_mbtn("👤 Личный кабинет", "menu_cabinet")])

    if tariff_type == "single":
        lines.extend(
            [
                "",
                "У Вас разовый тариф. Пожалуйста, направьте чек об оплате следующего занятия.",
            ]
        )
        if reply_markup is None:
            reply_markup = build_payment_prompt_keyboard_clean()
        if max_kb is None:
            from shared.max_api import btn as _mbtn, keyboard as _mkb
            max_kb = _mkb([_mbtn("💳 Оплатить занятия", "menu_paid")])

    if max_kb is None:
        from shared.max_api import btn as _mbtn, keyboard as _mkb
        max_kb = _mkb([_mbtn("👤 Личный кабинет", "menu_cabinet")])

    text = "\n".join(lines)
    if student_telegram_id:
        await send_student_notification(callback, student_telegram_id, text, reply_markup=reply_markup)
    if student_max_id:
        await _send_max_notification(student_max_id, text, max_kb)


def format_debt_report_text(report_data: dict, overdue_days: int) -> str:
    report_date = report_data.get("report_date", "-")
    total_current_debts = report_data.get("total_current_debts", 0)
    new_debts = report_data.get("new_debts", [])
    closed_debts = report_data.get("closed_debts", [])
    overdue_debts = report_data.get("overdue_debts", [])

    lines = [
        f"📊 <b>Отчёт по долгам за {report_date}</b>",
        "",
        f"Текущих долгов по направлениям: <b>{total_current_debts}</b>",
        f"Новые долги за день: <b>{len(new_debts)}</b>",
        f"Закрытые долги за день: <b>{len(closed_debts)}</b>",
        f"Долги старше {overdue_days} дн.: <b>{len(overdue_debts)}</b>",
    ]

    if new_debts:
        lines.append("")
        lines.append("<b>Новые долги:</b>")
        for item in new_debts[:20]:
            lines.append(
                f"- {item.get('student_name', '—')} | {item.get('subject_name', '—')} — "
                f"{item.get('teacher_name', '—')} | долг: {abs(item.get('lesson_balance', 0))}"
            )

    if closed_debts:
        lines.append("")
        lines.append("<b>Закрытые долги:</b>")
        for item in closed_debts[:20]:
            lines.append(
                f"- {item.get('student_name', '—')} | {item.get('subject_name', '—')} — "
                f"{item.get('teacher_name', '—')}"
            )

    if overdue_debts:
        lines.append("")
        lines.append(f"<b>Не оплатили более {overdue_days} дней:</b>")
        for item in overdue_debts[:30]:
            lines.append(
                f"- {item.get('student_name', '—')} | {item.get('subject_name', '—')} — "
                f"{item.get('teacher_name', '—')} | дней: {item.get('age_days', 0)} | "
                f"долг: {abs(item.get('lesson_balance', 0))}"
            )

    return "\n".join(lines)


def get_debtors_keyboard(debtors: list[dict]) -> InlineKeyboardMarkup:
    buttons: list[list[InlineKeyboardButton]] = []
    for debtor in debtors[:40]:
        student_id = int(debtor["student_id"])
        full_name = str(debtor.get("full_name") or f"Ученик #{student_id}")
        username = debtor.get("telegram_username")
        total_debt = int(debtor.get("total_debt_lessons") or 0)
        suffix = f"@{username}" if username else f"ID:{student_id}"
        text = f"{full_name} | {suffix} | долг: {total_debt}"
        buttons.append(
            [InlineKeyboardButton(text=text[:64], callback_data=f"admin_debtor_{student_id}")]
        )

    buttons.append([InlineKeyboardButton(text="Найти ученика / переименовать", callback_data="admin_find_student")])
    buttons.append([InlineKeyboardButton(text="Главное меню", callback_data="menu_home")])
    return InlineKeyboardMarkup(inline_keyboard=buttons)


def get_debtor_details_keyboard(telegram_id: int | None, username: str | None) -> InlineKeyboardMarkup:
    buttons: list[list[InlineKeyboardButton]] = []
    if telegram_id:
        buttons.append([InlineKeyboardButton(text="Открыть чат в Telegram", url=f"tg://user?id={telegram_id}")])
    elif username:
        buttons.append([InlineKeyboardButton(text="Открыть профиль", url=f"https://t.me/{username}")])
    buttons.append([InlineKeyboardButton(text="← К списку должников", callback_data="admin_debtors")])
    buttons.append([InlineKeyboardButton(text="Главное меню", callback_data="menu_home")])
    return InlineKeyboardMarkup(inline_keyboard=buttons)


def get_student_contact_keyboard(telegram_id: int | None, username: str | None) -> InlineKeyboardMarkup | None:
    buttons: list[list[InlineKeyboardButton]] = []
    safe_username = (username or "").strip().lstrip("@")
    if telegram_id:
        buttons.append([InlineKeyboardButton(text="Открыть чат в Telegram", url=f"tg://user?id={telegram_id}")])
    elif safe_username:
        buttons.append([InlineKeyboardButton(text="Открыть профиль в Telegram", url=f"https://t.me/{safe_username}")])
    if not buttons:
        return None
    return InlineKeyboardMarkup(inline_keyboard=buttons)


async def send_student_contact_shortcut(
    target: Message,
    *,
    telegram_id: int | None,
    username: str | None,
) -> None:
    contact_keyboard = get_student_contact_keyboard(telegram_id, username)
    if not contact_keyboard:
        return
    try:
        await target.answer("Быстрый переход в чат с учеником:", reply_markup=contact_keyboard)
    except Exception as exc:
        logger.warning("Cannot send student contact shortcut: %s", exc)


def can_delete_role(actor_id: int, target_role: str) -> bool:
    if actor_id in SUPERADMINS:
        return target_role in {"admin", "teacher", "student"}
    return target_role in {"teacher", "student"}


def delete_user_with_related_data(target_role: str, target_telegram_id: int) -> dict:
    if target_role == "admin":
        return delete_admin_by_telegram_id(target_telegram_id)
    if target_role == "teacher":
        return delete_teacher_by_telegram_id(target_telegram_id)
    if target_role == "student":
        return delete_student_by_telegram_id(target_telegram_id)
    return {"ok": False}


def get_teacher_owned_directions(teacher_telegram_id: int, student_id: int):
    teacher = get_teacher_by_telegram_id(teacher_telegram_id)
    if not teacher:
        return []

    teacher_id = teacher[0]
    directions = get_student_directions(student_id)
    result = []

    for direction in directions:
        direction_id = direction[0]
        lesson = get_student_lesson_by_id(direction_id)
        if lesson and lesson[2] == teacher_id:
            result.append(direction)

    return result


def can_manage_attendance(user_id: int, direction_id: int) -> bool:
    role = get_role_by_user_id(user_id)
    if role in ["superadmin", "admin"]:
        return True

    if role != "teacher":
        return False

    teacher = get_teacher_by_telegram_id(user_id)
    lesson = get_student_lesson_by_id(direction_id)
    if not teacher or not lesson:
        return False

    return lesson[2] == teacher[0]


def load_teacher_names_for_binding() -> list[str]:
    names: list[str] = []
    for _teacher_id, full_name, _subject_name, _username in list_teacher_profiles(limit=2000):
        if full_name and full_name not in names:
            names.append(full_name)

    return names


def is_valid_username(value: str) -> bool:
    return bool(re.fullmatch(r"@[A-Za-z0-9_]{5,32}", value.strip()))


def build_onboarding_link(token: str) -> str | None:
    if not SCHOOL_BOT_USERNAME:
        return None
    return f"https://t.me/{SCHOOL_BOT_USERNAME}?start=invite_{token}"


def parse_publication_links(raw_text: str) -> list[str]:
    value = (raw_text or "").strip()
    if not value or value == "-":
        return []

    links: list[str] = []
    for token in re.split(r"[\s,;]+", value):
        token = token.strip()
        if not token:
            continue
        if token.startswith("@"):
            token = f"https://t.me/{token.lstrip('@')}"
        if token.startswith("http://") or token.startswith("https://"):
            if token not in links:
                links.append(token)
    return links[:8]


def parse_publication_schedule(value: str) -> datetime | None:
    text = value.strip()
    for fmt in ("%d.%m.%Y %H:%M", "%Y-%m-%d %H:%M"):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    return None


def build_links_block(links: list[str]) -> str:
    if not links:
        return ""
    lines = ["", "<b>Ссылки:</b>"]
    for idx, link in enumerate(links, start=1):
        lines.append(f"{idx}. {link}")
    return "\n".join(lines)


@router.message(CommandStart())
async def start_handler(message: Message, state: FSMContext):
    await state.clear()
    user_id = message.from_user.id
    username = normalize_telegram_username(message.from_user.username)

    upsert_known_telegram_user(
        telegram_id=user_id,
        telegram_username=username,
        full_name=message.from_user.full_name,
    )

    if user_id in SUPERADMINS:
        existing_user = get_user_by_telegram_id(user_id)
        if not existing_user:
            add_user(
                telegram_id=user_id,
                full_name=message.from_user.full_name,
                role="superadmin"
            )

        await message.answer(
            "🔐 <b>Главное меню супер-администратора</b>\n\nДобро пожаловать во внутренний бот школы. Выберите раздел:",
            reply_markup=get_superadmin_menu(),
            parse_mode="HTML"
        )
        return

    user = get_user_by_telegram_id(user_id)

    if not user:
        pending_admin_invite = get_latest_pending_invite_by_role_and_username("admin", username)
        if pending_admin_invite:
            invite_id, _token, _role, invite_full_name, _invite_username, _entity_type, _entity_id = pending_admin_invite
            add_user(
                telegram_id=user_id,
                full_name=invite_full_name or message.from_user.full_name,
                role="admin",
                telegram_username=username,
            )
            mark_onboarding_invite_used(invite_id=int(invite_id), telegram_id=user_id)
            await message.answer(
                "✅ <b>Доступ активирован</b>\n\n"
                "⚙️ <b>Админ-панель</b>\n\nДобро пожаловать во внутренний бот школы. Выберите раздел:",
                reply_markup=get_admin_menu(),
                parse_mode="HTML"
            )
            return

        await message.answer("У тебя нет доступа к этому боту.")
        return

    _, telegram_id, full_name, role, is_active = user

    if not is_active:
        await message.answer("Твой доступ отключен.")
        return

    if role == "admin":
        await message.answer(
            "⚙️ <b>Админ-панель</b>\n\nДобро пожаловать во внутренний бот школы. Выберите раздел:",
            reply_markup=get_admin_menu(),
            parse_mode="HTML"
        )
        return

    if role == "teacher":
        await message.answer(
            "Внутренний бот школы.\n\nТы вошел как преподаватель.",
            reply_markup=get_teacher_menu()
        )
        return

    if role == "student":
        await message.answer(
            "Личный кабинет ученика.",
            reply_markup=get_student_menu()
        )
        return

    await message.answer("У тебя нет доступа к этому боту.")


@router.message(Command("menu"))
async def menu_handler(message: Message, state: FSMContext):
    await start_handler(message, state)


@router.callback_query(lambda c: c.data == "menu_home")
async def menu_home(callback: CallbackQuery, state: FSMContext):
    await state.clear()
    menu = get_home_menu_by_user_id(callback.from_user.id)
    if menu is None:
        await callback.message.answer("Доступ не найден. Используйте /start для повторного входа.")
        await callback.answer()
        return
    # Show different messages for different user types
    if callback.from_user.id in SUPERADMINS:
        message_text = "🔐 <b>Главное меню супер-администратора</b>\n\nВыберите раздел:"
    elif is_admin_role(callback.from_user.id):
        message_text = "⚙️ <b>Админ-панель</b>\n\nВыберите раздел:"
    else:
        message_text = "📋 <b>Главное меню</b>\n\nВыберите действие:"
    await callback.message.answer(message_text, reply_markup=menu, parse_mode="HTML")
    await callback.answer()


@router.callback_query(lambda c: c.data == "superadmin_section_users")
async def superadmin_section_users(callback: CallbackQuery):
    if callback.from_user.id not in SUPERADMINS:
        await callback.answer("Нет доступа", show_alert=True)
        return
    await callback.message.answer("👥 <b>Управление пользователями</b>\n\nВыберите действие:", reply_markup=get_superadmin_users_menu(), parse_mode="HTML")
    await callback.answer()


@router.callback_query(lambda c: c.data == "superadmin_section_school")
async def superadmin_section_school(callback: CallbackQuery):
    if callback.from_user.id not in SUPERADMINS:
        await callback.answer("Нет доступа", show_alert=True)
        return
    await callback.message.answer("📚 <b>Учебный процесс</b>\n\nВыберите действие:", reply_markup=get_superadmin_school_menu(), parse_mode="HTML")
    await callback.answer()


@router.callback_query(lambda c: c.data == "superadmin_section_reports")
async def superadmin_section_reports(callback: CallbackQuery):
    if callback.from_user.id not in SUPERADMINS:
        await callback.answer("Нет доступа", show_alert=True)
        return
    await callback.message.answer("📊 <b>Отчеты и журнал</b>\n\nВыберите действие:", reply_markup=get_superadmin_reports_menu(), parse_mode="HTML")
    await callback.answer()


@router.callback_query(lambda c: c.data == "superadmin_back_main")
async def superadmin_back_main(callback: CallbackQuery):
    if callback.from_user.id not in SUPERADMINS:
        await callback.answer("Нет доступа", show_alert=True)
        return
    await callback.message.answer("🔐 <b>Главное меню супер-администратора</b>\n\nВыберите раздел:", reply_markup=get_superadmin_menu(), parse_mode="HTML")
    await callback.answer()


@router.callback_query(lambda c: c.data == "admin_dashboard")
async def admin_dashboard(callback: CallbackQuery):
    if not is_admin_role(callback.from_user.id):
        await callback.answer("Нет доступа", show_alert=True)
        return
    metrics = get_admin_dashboard_metrics()

    pending_total = metrics["payments_pending"] + metrics["payments_processing"]

    lines = [
        "📊 <b>ДАШБОРД</b>",
        "",
        "<b>💳 Оплаты</b>",
        f"   • На проверке: <b>{pending_total}</b>"
        f"  (ожидают: {metrics['payments_pending']},"
        f" в обработке: {metrics['payments_processing']})",
        f"   • Просрочено за 7 дней: <b>{metrics['payments_expired_week']}</b>",
        "",
        "<b>📕 Долги</b>",
        f"   • Должников: <b>{metrics['debtors_count']}</b>",
        f"   • Всего занятий долга: <b>{metrics['debt_lessons_total']}</b>",
        "",
        "<b>🎓 Активность за 7 дней</b>",
        f"   • Проведено занятий: <b>{metrics['lessons_attended_week']}</b>",
        f"   • Новых учеников: <b>{metrics['new_students_week']}</b>",
        "",
        "<b>🤝 Реферальная программа</b>",
        f"   • Зафиксировано приглашений: <b>{metrics['referrals_captured']}</b>",
        f"   • Связано с карточкой ученика: <b>{metrics['referrals_linked']}</b>",
        f"   • Награждено бонусом: <b>{metrics['referrals_rewarded']}</b>",
    ]

    await callback.message.answer("\n".join(lines), parse_mode="HTML")
    await callback.answer()


@router.callback_query(lambda c: c.data == "admin_section_management")
async def admin_section_management(callback: CallbackQuery):
    if not is_admin_role(callback.from_user.id):
        await callback.answer("Нет доступа", show_alert=True)
        return
    await callback.message.answer("👥 <b>Управление пользователями</b>\n\nВыберите действие:", reply_markup=get_admin_management_menu(), parse_mode="HTML")
    await callback.answer()


@router.callback_query(lambda c: c.data == "admin_section_education")
async def admin_section_education(callback: CallbackQuery):
    if not is_admin_role(callback.from_user.id):
        await callback.answer("Нет доступа", show_alert=True)
        return
    await callback.message.answer("📚 <b>Учеба и занятия</b>\n\nВыберите действие:", reply_markup=get_admin_education_menu(), parse_mode="HTML")
    await callback.answer()


@router.callback_query(lambda c: c.data == "admin_section_finance")
async def admin_section_finance(callback: CallbackQuery):
    if not is_admin_role(callback.from_user.id):
        await callback.answer("Нет доступа", show_alert=True)
        return
    await callback.message.answer("💰 <b>Финансы и баланс</b>\n\nВыберите действие:", reply_markup=get_admin_finance_menu(), parse_mode="HTML")
    await callback.answer()


@router.callback_query(lambda c: c.data == "admin_section_reports")
async def admin_section_reports(callback: CallbackQuery):
    if not is_admin_role(callback.from_user.id):
        await callback.answer("Нет доступа", show_alert=True)
        return
    await callback.message.answer("📊 <b>Отчеты и аналитика</b>\n\nВыберите действие:", reply_markup=get_admin_reports_menu(), parse_mode="HTML")
    await callback.answer()


@router.callback_query(lambda c: c.data == "admin_back_main")
async def admin_back_main(callback: CallbackQuery):
    if not is_admin_role(callback.from_user.id):
        await callback.answer("Нет доступа", show_alert=True)
        return
    await callback.message.answer("⚙️ <b>Админ-панель</b>\n\nВыберите раздел:", reply_markup=get_admin_menu(), parse_mode="HTML")
    await callback.answer()


@router.callback_query(lambda c: c.data == "admin_publication_new")
async def admin_publication_new(callback: CallbackQuery, state: FSMContext):
    if not is_admin_role(callback.from_user.id):
        await callback.answer("Нет доступа", show_alert=True)
        return

    await state.clear()
    await callback.message.answer(
        "Введите текст публикации для учеников.\n"
        "Описание обязательно.",
        reply_markup=get_main_menu_shortcut_keyboard(),
    )
    await state.set_state(AdminStates.waiting_publication_description)
    await callback.answer()


@router.callback_query(lambda c: c.data == "admin_payment_chat_message")
async def admin_payment_chat_message(callback: CallbackQuery, state: FSMContext):
    if not is_admin_role(callback.from_user.id):
        await callback.answer("Нет доступа", show_alert=True)
        return

    await state.clear()
    await callback.message.answer(
        "Введите сообщение для чата оплат.\n"
        "Оно будет отправлено от имени бота в чат оплат.",
        reply_markup=get_main_menu_shortcut_keyboard(),
    )
    await state.set_state(AdminStates.waiting_payment_chat_message)
    await callback.answer()


@router.message(AdminStates.waiting_payment_chat_message)
async def admin_payment_chat_message_send(message: Message, state: FSMContext):
    if not is_admin_role(message.from_user.id):
        await message.answer("Нет доступа.")
        await state.clear()
        return

    if not SCHOOL_BOT_PAYMENTS_CHAT_ID:
        await message.answer(
            "Не задан SCHOOL_BOT_PAYMENTS_CHAT_ID в .env.",
            reply_markup=get_admin_reply_menu(message.from_user.id),
        )
        await state.clear()
        return

    text = (message.text or "").strip()
    if not text:
        await message.answer("Введите текст сообщения.")
        return

    sender_name = message.from_user.full_name or "Администратор"
    payload = f"Сообщение от администратора {sender_name}:\n\n{text}"

    try:
        if SCHOOL_BOT_TOKEN:
            async with Bot(token=SCHOOL_BOT_TOKEN) as school_bot:
                await school_bot.send_message(SCHOOL_BOT_PAYMENTS_CHAT_ID, payload)
        else:
            await message.bot.send_message(SCHOOL_BOT_PAYMENTS_CHAT_ID, payload)
    except Exception as exc:
        await message.answer(
            f"Не удалось отправить сообщение в чат оплат: {exc}",
            reply_markup=get_admin_reply_menu(message.from_user.id),
        )
        await state.clear()
        return

    log_admin_action(
        admin_telegram_id=message.from_user.id,
        action_type="manual_message_to_payments_chat",
        target_type="chat",
        target_id=SCHOOL_BOT_PAYMENTS_CHAT_ID,
        details={"text_preview": text[:200]},
        status="success",
    )
    await message.answer(
        "Сообщение отправлено в чат оплат.",
        reply_markup=get_admin_reply_menu(message.from_user.id),
    )
    await state.clear()


@router.message(AdminStates.waiting_publication_description)
async def admin_publication_description(message: Message, state: FSMContext):
    text = (message.text or "").strip()
    if len(text) < 3:
        await message.answer("Описание слишком короткое. Введите полноценный текст публикации.")
        return

    await state.update_data(publication_description=text)
    await message.answer(
        "Теперь отправьте фото для публикации.\n"
        "Если фото не нужно, отправьте символ: -"
    )
    await state.set_state(AdminStates.waiting_publication_photo)


@router.message(AdminStates.waiting_publication_photo)
async def admin_publication_photo(message: Message, state: FSMContext):
    photo_file_id = None
    text_value = (message.text or "").strip()

    if message.photo:
        photo_file_id = message.photo[-1].file_id
    elif text_value != "-":
        await message.answer("Отправьте фото или символ - чтобы пропустить.")
        return

    await state.update_data(publication_photo_file_id=photo_file_id)
    await message.answer(
        "Добавьте ссылки (через пробел или с новой строки).\n"
        "Можно вставлять URL или @username.\n"
        "Если ссылки не нужны, отправьте: -"
    )
    await state.set_state(AdminStates.waiting_publication_links)


@router.message(AdminStates.waiting_publication_links)
async def admin_publication_links(message: Message, state: FSMContext):
    raw = (message.text or "").strip()
    links = parse_publication_links(raw)
    if raw and raw != "-" and not links:
        await message.answer(
            "Не удалось распознать ссылки.\n"
            "Используйте формат https://... или @username, либо отправьте -."
        )
        return

    await state.update_data(publication_links=links)
    await message.answer(
        "Выберите аудиторию публикации:",
        reply_markup=get_publication_audience_keyboard(),
    )
    await state.set_state(AdminStates.waiting_publication_audience)


@router.callback_query(
    AdminStates.waiting_publication_audience,
    lambda c: c.data in {
        "publication_audience_students",
        "publication_audience_students_plus_me",
        "publication_audience_me_only",
    },
)
async def admin_publication_audience(callback: CallbackQuery, state: FSMContext):
    if not is_admin_role(callback.from_user.id):
        await callback.answer("Нет доступа", show_alert=True)
        return

    audience_map = {
        "publication_audience_students": "students",
        "publication_audience_students_plus_me": "students_plus_creator",
        "publication_audience_me_only": "creator_only",
    }
    audience = audience_map.get(callback.data, "students")
    await state.update_data(publication_audience=audience)
    await callback.message.answer(
        "Выберите, когда отправить публикацию:",
        reply_markup=get_publication_schedule_keyboard(),
    )
    await state.set_state(AdminStates.waiting_publication_schedule_mode)
    await callback.answer()


@router.callback_query(
    AdminStates.waiting_publication_schedule_mode,
    lambda c: c.data in {"publication_send_now", "publication_schedule_pick_time"},
)
async def admin_publication_schedule_mode(callback: CallbackQuery, state: FSMContext):
    if not is_admin_role(callback.from_user.id):
        await callback.answer("Нет доступа", show_alert=True)
        return

    data = await state.get_data()
    description = (data.get("publication_description") or "").strip()
    photo_file_id = data.get("publication_photo_file_id")
    links = data.get("publication_links") or []
    audience = data.get("publication_audience") or "students"

    if not description:
        await state.clear()
        await callback.message.answer("Сценарий публикации сброшен. Начните заново.")
        await callback.answer()
        return

    if callback.data == "publication_send_now":
        scheduled_for = msk_now_naive().strftime("%Y-%m-%d %H:%M:%S")
        post_id = create_publication_post(
            created_by=callback.from_user.id,
            audience=audience,
            description=description,
            photo_file_id=photo_file_id,
            links=links,
            scheduled_for=scheduled_for,
        )
        log_admin_action(
            admin_telegram_id=callback.from_user.id,
            action_type="publication_created",
            target_type="publication_post",
            target_id=post_id,
            details={
                "mode": "now",
                "audience": audience,
                "has_photo": bool(photo_file_id),
                "links_count": len(links),
            },
            status="success",
        )
        await state.clear()
        await callback.message.answer(
            f"Публикация создана и поставлена в очередь отправки (ID: {post_id}, время МСК).",
            reply_markup=get_admin_reply_menu(callback.from_user.id),
        )
        await callback.answer("Готово")
        return

    await callback.message.answer(
        "Введите дату и время публикации.\n"
        "Время указывайте по МСК.\n"
        "Формат: ДД.ММ.ГГГГ ЧЧ:ММ\n"
        "Например: 25.04.2026 10:30"
    )
    await state.set_state(AdminStates.waiting_publication_schedule_datetime)
    await callback.answer()


@router.message(AdminStates.waiting_publication_schedule_datetime)
async def admin_publication_schedule_datetime(message: Message, state: FSMContext):
    if not is_admin_role(message.from_user.id):
        await message.answer("Нет доступа.")
        await state.clear()
        return

    schedule_dt = parse_publication_schedule(message.text or "")
    if schedule_dt is None:
        await message.answer("Неверный формат даты. Укажите время по МСК. Пример: 25.04.2026 10:30")
        return
    if schedule_dt <= msk_now_naive():
        await message.answer("Дата должна быть в будущем (по МСК). Укажите более позднее время.")
        return

    data = await state.get_data()
    description = (data.get("publication_description") or "").strip()
    if not description:
        await state.clear()
        await message.answer("Сценарий публикации сброшен. Начните заново.")
        return

    photo_file_id = data.get("publication_photo_file_id")
    links = data.get("publication_links") or []
    audience = data.get("publication_audience") or "students"
    scheduled_for = schedule_dt.strftime("%Y-%m-%d %H:%M:%S")

    post_id = create_publication_post(
        created_by=message.from_user.id,
        audience=audience,
        description=description,
        photo_file_id=photo_file_id,
        links=links,
        scheduled_for=scheduled_for,
    )
    log_admin_action(
        admin_telegram_id=message.from_user.id,
        action_type="publication_created",
        target_type="publication_post",
        target_id=post_id,
        details={
            "mode": "scheduled",
            "audience": audience,
            "scheduled_for": scheduled_for,
            "has_photo": bool(photo_file_id),
            "links_count": len(links),
        },
        status="success",
    )

    await state.clear()
    await message.answer(
        f"Публикация запланирована на {schedule_dt.strftime('%d.%m.%Y %H:%M')} МСК (ID: {post_id}).",
        reply_markup=get_admin_reply_menu(message.from_user.id),
    )


@router.callback_query(lambda c: c.data == "admin_review_new")
async def admin_review_new(callback: CallbackQuery, state: FSMContext):
    if not is_admin_role(callback.from_user.id):
        await callback.answer("Нет доступа", show_alert=True)
        return

    await state.clear()
    await callback.message.answer(
        "Введите текст карточки отзыва.\n"
        "Описание обязательное.",
        reply_markup=get_main_menu_shortcut_keyboard(),
    )
    await state.set_state(AdminStates.waiting_review_description)
    await callback.answer()


@router.message(AdminStates.waiting_review_description)
async def admin_review_description(message: Message, state: FSMContext):
    text = (message.text or "").strip()
    if len(text) < 3:
        await message.answer("Описание слишком короткое. Введите более подробный текст.")
        return

    await state.update_data(review_description=text)
    await message.answer(
        "Теперь отправьте фото или файл (pdf/doc), если нужно.\n"
        "Если медиа не требуется, отправьте символ: -"
    )
    await state.set_state(AdminStates.waiting_review_media)


@router.message(AdminStates.waiting_review_media)
async def admin_review_media(message: Message, state: FSMContext, bot: Bot):
    media_file_id = None
    media_type = None
    media_local_path = None
    text_value = (message.text or "").strip().lower()

    # Ensure reviews directory exists
    reviews_dir = Path("assets/reviews")
    reviews_dir.mkdir(parents=True, exist_ok=True)

    if message.photo:
        photo = message.photo[-1]
        media_file_id = photo.file_id
        media_type = "photo"
        media_local_path = None

        # Download and save photo
        try:
            timestamp = int(time.time())
            filename = f"review_{timestamp}_{uuid4().hex[:8]}.jpg"
            target_path = reviews_dir / filename

            print(f"📥 Downloading photo...")
            print(f"   Target: {target_path}")
            print(f"   File ID: {media_file_id[:30]}...")

            # Get file from Telegram
            file = await bot.get_file(media_file_id)
            print(f"   File path on Telegram: {file.file_path}")

            # Download file using correct aiogram method
            await bot.download(file, destination=target_path)

            if target_path.exists():
                # Store as relative path for portability
                media_local_path = f"assets/reviews/{filename}"
                print(f"   ✅ Saved as: {media_local_path}")

                await message.answer(
                    f"✅ <b>Фото добавлено</b>\n"
                    f"📷 Фото сохранено локально\n"
                    f"🔗 Telegram ID: {media_file_id[:20]}...\n"
                    f"💾 Размер: {target_path.stat().st_size / 1024:.1f} KB",
                    parse_mode="HTML"
                )
            else:
                media_local_path = None
                await message.answer(
                    f"✅ <b>Фото добавлено</b>\n"
                    f"📷 Хранилище: Telegram\n"
                    f"⚠️ Локальное сохранение не работает",
                    parse_mode="HTML"
                )
        except Exception as e:
            print(f"❌ Error saving photo: {e}")
            import traceback
            traceback.print_exc()
            logging.error(f"Could not save photo locally: {e}")
            media_local_path = None
            await message.answer(
                f"✅ <b>Фото добавлено</b>\n"
                f"📷 Хранилище: Telegram\n"
                f"⚠️ Локальное сохранение не работает",
                parse_mode="HTML"
            )

    elif message.document:
        document = message.document
        media_file_id = document.file_id
        media_type = "document"
        media_local_path = None

        # Download and save document
        try:
            timestamp = int(time.time())
            file_ext = Path(document.file_name or "document.pdf").suffix
            filename = f"review_{timestamp}_{uuid4().hex[:8]}{file_ext}"
            target_path = reviews_dir / filename

            file = await bot.get_file(media_file_id)
            await bot.download(file, destination=target_path)

            if target_path.exists():
                # Store as relative path for portability
                media_local_path = f"assets/reviews/{filename}"

                await message.answer(
                    f"✅ <b>Файл добавлен</b>\n"
                    f"📄 {document.file_name or 'документ'}\n"
                    f"💾 Размер: {document.file_size / 1024:.1f} KB\n"
                    f"✅ Сохранено локально",
                    parse_mode="HTML"
                )
            else:
                media_local_path = None
                await message.answer(
                    f"✅ <b>Файл добавлен</b>\n"
                    f"📄 {document.file_name or 'документ'}\n"
                    f"💾 Размер: {document.file_size / 1024:.1f} KB\n"
                    f"⚠️ Хранилище: Telegram",
                    parse_mode="HTML"
                )
        except Exception as e:
            # If local save fails, continue with Telegram file_id
            logging.warning(f"Could not save document locally: {e}")
            media_local_path = None
            await message.answer(
                f"✅ <b>Файл добавлен</b>\n"
                f"📄 {document.file_name or 'документ'}\n"
                f"💾 Размер: {document.file_size / 1024:.1f} KB\n"
                f"⚠️ Хранилище: Telegram",
                parse_mode="HTML"
            )

    elif text_value == "-" or text_value == "нет":
        media_file_id = None
        media_type = None
        media_local_path = None
        await message.answer("⏭️ <b>Отзыв без медиа</b>\n\n📝 Продолжаем добавлять описание", parse_mode="HTML")
    else:
        await message.answer(
            "❌ <b>Неверный формат!</b>\n\n"
            "Пожалуйста, отправьте:\n"
            "📷 <b>Фото</b> (JPG, PNG) ИЛИ\n"
            "📄 <b>Файл</b> (PDF, DOC и т.д.) ИЛИ\n"
            "➖ Введите: <code>-</code> (чтобы пропустить)",
            parse_mode="HTML"
        )
        return

    await state.update_data(
        review_media_file_id=media_file_id,
        review_media_type=media_type,
        review_media_local_path=media_local_path
    )
    await message.answer(
        "🔗 <b>Добавьте ссылки</b> (если есть)\n\n"
        "Введите через пробел или новую строку:\n"
        "• URL: <code>https://example.com</code>\n"
        "• Username: <code>@myusername</code>\n\n"
        "Или отправьте <code>-</code> если ссылок нет",
        parse_mode="HTML"
    )
    await state.set_state(AdminStates.waiting_review_links)


@router.message(AdminStates.waiting_review_links)
async def admin_review_links(message: Message, state: FSMContext):
    if not is_admin_role(message.from_user.id):
        await message.answer("Нет доступа.")
        await state.clear()
        return

    raw = (message.text or "").strip()
    links = parse_publication_links(raw)
    if raw and raw != "-" and not links:
        await message.answer(
            "Не удалось распознать ссылки.\n"
            "Используйте формат https://... или @username, либо отправьте -."
        )
        return

    data = await state.get_data()
    description = (data.get("review_description") or "").strip()
    media_file_id = data.get("review_media_file_id")
    media_type = data.get("review_media_type")
    media_local_path = data.get("review_media_local_path")

    if not description:
        await state.clear()
        await message.answer("❌ <b>Сценарий отменён</b>\n\nОписание отзыва не заполнено. Пожалуйста, запустите снова.", parse_mode="HTML")
        return

    # Info message with emoji
    media_info = ""
    if media_type == "photo":
        media_info = f"📷 Фото (сохранено локально)"
    elif media_type == "document":
        media_info = f"📄 Документ (сохранено локально)"
    else:
        media_info = f"⏭️ Без медиа"

    links_info = f"🔗 Ссылок: {len(links)}" if links else "🔗 Ссылок: нет"

    info_msg = (
        f"📋 <b>Создание отзыва</b>\n\n"
        f"✍️ Описание: {len(description)} символов\n"
        f"{media_info}\n"
        f"{links_info}\n\n"
        f"⏳ Сохраняю в базу..."
    )
    await message.answer(info_msg, parse_mode="HTML")

    review_id = create_review_card(
        created_by=message.from_user.id,
        description=description,
        media_file_id=media_file_id,
        media_type=media_type,
        links=links,
        media_local_path=media_local_path,
    )
    log_admin_action(
        admin_telegram_id=message.from_user.id,
        action_type="review_card_created",
        target_type="review_card",
        target_id=review_id,
        details={
            "has_media": bool(media_file_id),
            "media_type": media_type,
            "links_count": len(links),
        },
        status="success",
    )

    caption = (
        f"<b>📝 Отзыв #{review_id}</b>\n\n"
        f"{description}"
        f"{build_links_block(links)}"
    )

    media_status = "✅"
    if media_type == "photo":
        media_status = "✅ 📷 С фото"
    elif media_type == "document":
        media_status = "✅ 📄 С документом"
    else:
        media_status = "✅ Текстовый отзыв"

    # Try to display the preview
    try:
        if media_local_path and media_type == "photo":
            from aiogram.types import FSInputFile
            # Convert relative path to absolute for FSInputFile
            abs_path = resolve_local_path(media_local_path)
            await message.answer_photo(
                photo=FSInputFile(abs_path),
                caption=caption,
                parse_mode="HTML"
            )
        elif media_local_path and media_type == "document":
            from aiogram.types import FSInputFile
            # Convert relative path to absolute for FSInputFile
            abs_path = resolve_local_path(media_local_path)
            await message.answer_document(
                document=FSInputFile(abs_path),
                caption=caption,
                parse_mode="HTML"
            )
        else:
            await message.answer(caption, parse_mode="HTML")
    except Exception as e:
        logging.error(f"Error displaying review preview: {e}")
        await message.answer(caption, parse_mode="HTML")

    await state.clear()
    await message.answer(
        f"<b>✅ Отзыв #{review_id} создан успешно!</b>\n\n"
        f"📊 Статус: {media_status}\n"
        f"📌 Количество ссылок: {len(links)}\n"
        f"👥 Видимость: Ученики в разделе ⭐ Отзывы\n\n"
        f"💡 <i>Отзыв сохранён с локальным хранилищем медиа</i>",
        parse_mode="HTML",
        reply_markup=get_admin_reply_menu(message.from_user.id),
    )


@router.callback_query(lambda c: c.data == "admin_review_list")
async def admin_review_list(callback: CallbackQuery):
    """Показать список всех отзывов с опцией удаления"""
    if not is_admin_role(callback.from_user.id):
        await callback.answer("Нет доступа", show_alert=True)
        return

    reviews = get_active_review_cards(limit=100)
    if not reviews:
        await callback.message.answer(
            "📭 <b>Отзывы не найдены</b>\n\n"
            "В системе нет активных отзывов. Создайте первый отзыв! ✨",
            parse_mode="HTML"
        )
        await callback.answer()
        return

    text = "<b>📋 Активные отзывы</b>\n\n"
    for idx, review in enumerate(reviews, 1):
        media_emoji = "📷" if review.get("media_type") == "photo" else "📄" if review.get("media_type") == "document" else "📝"
        desc_preview = (review.get("description") or "Нет описания")[:35]
        links_count = len(review.get("links") or [])
        links_emoji = f"🔗 {links_count}" if links_count > 0 else ""

        text += (
            f"<b>#{review['id']}</b> {media_emoji}\n"
            f"  📌 {desc_preview}{'...' if len(review.get('description', '')) > 35 else ''}\n"
            f"  {links_emoji}\n\n"
        )

    await callback.message.answer(text, parse_mode="HTML")

    # Кнопки управления отзывами
    buttons = []
    for review in reviews[:10]:
        media_icon = "📷" if review.get("media_type") == "photo" else "📄" if review.get("media_type") == "document" else "📝"
        rid = review['id']
        buttons.append([
            InlineKeyboardButton(text=f"📷 #{rid} фото", callback_data=f"admin_edit_review_photo_{rid}"),
            InlineKeyboardButton(text=f"🗑️ #{rid}", callback_data=f"admin_delete_review_{rid}"),
        ])

    if buttons:
        buttons.append([InlineKeyboardButton(text="↩️ Назад", callback_data="menu_home")])
        keyboard = InlineKeyboardMarkup(inline_keyboard=buttons)
        await callback.message.answer(
            "✏️ <b>Управление отзывами</b>\n\n"
            "📷 — заменить/добавить фото\n"
            "🗑️ — удалить отзыв",
            parse_mode="HTML",
            reply_markup=keyboard,
        )

    await callback.answer()


@router.callback_query(lambda c: c.data.startswith("admin_delete_review_"))
async def admin_delete_review(callback: CallbackQuery):
    """Удалить отзыв"""
    if not is_admin_role(callback.from_user.id):
        await callback.answer("Нет доступа", show_alert=True)
        return

    try:
        review_id = int(callback.data.split("_")[-1])
    except (ValueError, IndexError):
        await callback.answer("❌ Ошибка при обработке удаления", show_alert=True)
        return

    success = deactivate_review_card(review_id)
    if success:
        log_admin_action(
            admin_telegram_id=callback.from_user.id,
            action_type="review_card_deleted",
            target_type="review_card",
            target_id=review_id,
            details={},
            status="success",
        )
        await callback.message.answer(
            f"<b>🗑️ Отзыв удалён</b>\n\n"
            f"📋 ID: #{review_id}\n"
            f"❌ Статус: Деактивирован\n"
            f"👥 Видимость: Больше не видно ученикам\n\n"
            f"💾 Данные сохранены в архиве (не удалены)",
            parse_mode="HTML"
        )
    else:
        await callback.message.answer(
            f"<b>❌ Ошибка удаления</b>\n\n"
            f"Не удалось деактивировать отзыв #{review_id}\n"
            f"Попробуйте ещё раз или обратитесь к разработчику",
            parse_mode="HTML"
        )

    await callback.answer()


@router.callback_query(lambda c: c.data.startswith("admin_edit_review_photo_"))
async def admin_edit_review_photo_start(callback: CallbackQuery, state: FSMContext):
    if not is_admin_role(callback.from_user.id):
        await callback.answer("Нет доступа", show_alert=True)
        return
    try:
        review_id = int(callback.data.split("_")[-1])
    except (ValueError, IndexError):
        await callback.answer("Ошибка", show_alert=True)
        return
    await state.update_data(edit_review_id=review_id)
    await state.set_state(AdminStates.waiting_review_edit_media)
    await callback.message.answer(
        f"📷 <b>Замена фото отзыва #{review_id}</b>\n\n"
        "Отправьте новое фото или файл.\n"
        "Чтобы убрать медиа совсем — отправьте <code>-</code>.",
        parse_mode="HTML",
        reply_markup=get_main_menu_shortcut_keyboard(),
    )
    await callback.answer()


@router.message(AdminStates.waiting_review_edit_media)
async def admin_edit_review_photo_receive(message: Message, state: FSMContext, bot: Bot):
    data = await state.get_data()
    review_id = data.get("edit_review_id")
    if not review_id:
        await message.answer("Ошибка: ID отзыва не найден.")
        await state.clear()
        return

    reviews_dir = Path("assets/reviews")
    reviews_dir.mkdir(parents=True, exist_ok=True)

    media_file_id = None
    media_type = None
    media_local_path = None
    text_value = (message.text or "").strip()

    if text_value == "-":
        pass  # clear media
    elif message.photo:
        photo = message.photo[-1]
        media_file_id = photo.file_id
        media_type = "photo"
        try:
            from datetime import datetime as _dt
            from uuid import uuid4
            timestamp = _dt.now().strftime("%Y%m%d_%H%M%S")
            filename = f"review_{timestamp}_{uuid4().hex[:8]}.jpg"
            target_path = reviews_dir / filename
            file_info = await bot.get_file(photo.file_id)
            await bot.download_file(file_info.file_path, destination=str(target_path))
            media_local_path = f"assets/reviews/{filename}"
            media_file_id = photo.file_unique_id
        except Exception:
            media_file_id = photo.file_id
    elif message.document:
        doc = message.document
        mime_type = (doc.mime_type or "").lower()
        file_name = (doc.file_name or "").lower()
        if mime_type != "application/pdf" and not file_name.endswith(".pdf"):
            await message.answer("❌ Поддерживаются только фото и PDF. Попробуйте ещё раз.")
            return
        media_file_id = doc.file_id
        media_type = "document"
        try:
            from datetime import datetime as _dt
            from uuid import uuid4
            ext = ".pdf"
            timestamp = _dt.now().strftime("%Y%m%d_%H%M%S")
            filename = f"review_{timestamp}_{uuid4().hex[:8]}{ext}"
            target_path = reviews_dir / filename
            file_info = await bot.get_file(doc.file_id)
            await bot.download_file(file_info.file_path, destination=str(target_path))
            media_local_path = f"assets/reviews/{filename}"
        except Exception:
            pass
    else:
        await message.answer("❌ Отправьте фото, PDF или «-» для удаления медиа.")
        return

    success = update_review_card_media(
        review_id,
        media_file_id=media_file_id,
        media_type=media_type,
        media_local_path=media_local_path,
    )
    await state.clear()
    if success:
        action = "удалено" if text_value == "-" else "обновлено"
        await message.answer(
            f"✅ <b>Медиа отзыва #{review_id} {action}</b>",
            parse_mode="HTML",
            reply_markup=get_admin_reply_menu(message.from_user.id),
        )
    else:
        await message.answer(
            f"❌ Отзыв #{review_id} не найден или уже удалён.",
            reply_markup=get_admin_reply_menu(message.from_user.id),
        )


@router.callback_query(lambda c: c.data == "admin_add_student")
async def admin_add_student(callback: CallbackQuery, state: FSMContext):
    if not is_admin_role(callback.from_user.id):
        await callback.answer("Нет доступа", show_alert=True)
        return

    await callback.message.answer(
        "Введите ФИО ученика.\nПодсказка: в любой момент можно нажать «Главное меню».",
        reply_markup=get_main_menu_shortcut_keyboard(),
    )
    await state.set_state(AdminStates.waiting_student_name)
    await callback.answer()


@router.callback_query(lambda c: c.data == "admin_bind_teacher_telegram")
async def admin_bind_teacher_telegram(callback: CallbackQuery, state: FSMContext):
    if not is_admin_role(callback.from_user.id):
        await callback.answer("Нет доступа", show_alert=True)
        return

    teacher_names = load_teacher_names_for_binding()
    if not teacher_names:
        await callback.message.answer("Не удалось получить список преподавателей.")
        await callback.answer()
        return

    await state.update_data(bind_teacher_names=teacher_names)
    await callback.message.answer(
        "Выбери преподавателя для привязки Telegram ID:",
        reply_markup=get_teacher_bind_keyboard(teacher_names)
    )
    await callback.answer()


@router.callback_query(lambda c: c.data == "admin_bind_teacher_cancel")
async def admin_bind_teacher_cancel(callback: CallbackQuery, state: FSMContext):
    if not is_admin_role(callback.from_user.id):
        await callback.answer("Нет доступа", show_alert=True)
        return

    await state.clear()
    await callback.message.answer("Привязка отменена.", reply_markup=get_admin_reply_menu(callback.from_user.id))
    await callback.answer()


@router.callback_query(lambda c: c.data.startswith("bind_teacher_choose_"))
async def choose_teacher_for_binding(callback: CallbackQuery, state: FSMContext):
    if not is_admin_role(callback.from_user.id):
        await callback.answer("Нет доступа", show_alert=True)
        return

    data = await state.get_data()
    teacher_names = data.get("bind_teacher_names") or load_teacher_names_for_binding()

    try:
        idx = int(callback.data.split("_")[-1])
    except ValueError:
        await callback.answer("Некорректный выбор", show_alert=True)
        return

    if idx < 0 or idx >= len(teacher_names):
        await callback.answer("Преподаватель не найден", show_alert=True)
        return

    teacher_name = teacher_names[idx]
    await state.update_data(bind_teacher_name=teacher_name, bind_teacher_names=teacher_names)

    await callback.message.answer(
        f"Выбран преподаватель: {teacher_name}\n\n"
        "Отправь Telegram ID преподавателя числом.\n"
        "Для отмены напиши: отмена"
    )
    await state.set_state(AdminStates.waiting_bind_teacher_telegram_id)
    await callback.answer()


@router.message(AdminStates.waiting_bind_teacher_telegram_id)
async def process_bind_teacher_telegram_id(message: Message, state: FSMContext):
    if not is_admin_role(message.from_user.id):
        await message.answer("Нет доступа.")
        await state.clear()
        return

    text = message.text.strip()
    if text.lower() in {"отмена", "-"}:
        await state.clear()
        await message.answer("Привязка отменена.", reply_markup=get_admin_reply_menu(message.from_user.id))
        return

    if not text.isdigit():
        await message.answer("Telegram ID должен быть числом. Для отмены напиши: отмена")
        return

    data = await state.get_data()
    teacher_name = data.get("bind_teacher_name")
    if not teacher_name:
        await state.clear()
        await message.answer(
            "Не удалось определить преподавателя. Начни заново через меню.",
            reply_markup=get_admin_reply_menu(message.from_user.id)
        )
        return

    telegram_id = int(text)
    result = bind_teacher_telegram_id(teacher_name, telegram_id)

    if not result["ok"]:
        log_admin_action(
            admin_telegram_id=message.from_user.id,
            action_type="bind_teacher_telegram_failed",
            target_type="teacher",
            target_id=None,
            details={"teacher_name": teacher_name, "error": result["error"]},
            status="error",
        )
        await message.answer(f"❌ {result['error']}\nПопробуй другой Telegram ID.")
        return

    add_user(
        telegram_id=telegram_id,
        full_name=teacher_name,
        role="teacher"
    )

    await message.answer(
        "✅ Преподаватель привязан.\n\n"
        f"Преподаватель: {teacher_name}\n"
        f"Telegram ID: {telegram_id}",
        reply_markup=get_admin_reply_menu(message.from_user.id)
    )
    log_admin_action(
        admin_telegram_id=message.from_user.id,
        action_type="bind_teacher_telegram_success",
        target_type="teacher",
        target_id=result.get("teacher_id"),
        details={
            "before": None,
            "after": {"teacher_name": teacher_name, "telegram_id": telegram_id},
        },
        status="success",
    )
    await state.clear()


@router.message(AdminStates.waiting_student_name)
async def get_student_name(message: Message, state: FSMContext):
    if not is_admin_role(message.from_user.id):
        await message.answer("Нет доступа.")
        await state.clear()
        return

    await state.update_data(full_name=message.text.strip())
    await message.answer(
        "Укажите @username ученика (обязательно), например: @ivan_ivanov"
    )
    await state.set_state(AdminStates.waiting_student_username)


@router.message(AdminStates.waiting_student_username)
async def get_student_username(message: Message, state: FSMContext):
    if not is_admin_role(message.from_user.id):
        await message.answer("Нет доступа.")
        await state.clear()
        return

    text = (message.text or "").strip()
    if not is_valid_username(text):
        await message.answer("Введите корректный @username в формате @example_user")
        return

    normalized_username = normalize_telegram_username(text)
    telegram_id = get_known_telegram_user_id_by_username(normalized_username)
    await state.update_data(
        telegram_id=telegram_id,
        telegram_username=normalized_username,
    )
    await message.answer("Введи номер телефона ученика или напиши '-' если его нет:")
    await state.set_state(AdminStates.waiting_student_phone)


@router.message(AdminStates.waiting_student_phone)
async def get_student_phone(message: Message, state: FSMContext):
    if not is_admin_role(message.from_user.id):
        await message.answer("Нет доступа.")
        await state.clear()
        return

    text = message.text.strip()
    phone = None if text == "-" else text

    data = await state.get_data()

    student_id = add_student(
        full_name=data["full_name"],
        telegram_id=data["telegram_id"],
        phone=phone,
        telegram_username=data.get("telegram_username"),
    )
    log_admin_action(
        admin_telegram_id=message.from_user.id,
        action_type="add_student",
        target_type="student",
        target_id=None,
        details={
            "before": None,
            "after": {
                "full_name": data["full_name"],
                "telegram_id": data["telegram_id"],
                "telegram_username": data.get("telegram_username"),
                "phone": phone,
            },
        },
        status="success",
    )

    if data["telegram_id"]:
        add_user(
            telegram_id=data["telegram_id"],
            full_name=data["full_name"],
            role="student",
            telegram_username=data.get("telegram_username"),
        )

    referral_text = ""
    if data["telegram_id"]:
        if link_invitee_student(int(data["telegram_id"]), int(student_id)):
            referral_row = get_referral_by_invitee_telegram_id(int(data["telegram_id"]))
            inviter_tg = referral_row[1] if referral_row else None
            referral_text = (
                "\n\n🎁 Ученик пришёл по реферальной ссылке"
                f" (приглашён tg_id={inviter_tg}). Ему будет применена скидка 20%"
                " на первое платное занятие, пригласивший получит бонус после оплаты."
            )

    onboarding_text = ""
    if not data["telegram_id"]:
        token = create_onboarding_invite(
            role="student",
            full_name=data["full_name"],
            telegram_username=data.get("telegram_username") or "",
            entity_type="student",
            entity_id=student_id,
            created_by=message.from_user.id,
        )
        link = build_onboarding_link(token)
        if link:
            onboarding_text = (
                "\n\nПользователь еще не писал боту, поэтому ID пока не найден.\n"
                "Отправьте ему эту ссылку для автоматической привязки:\n"
                f"{link}"
            )
        else:
            onboarding_text = (
                "\n\nПользователь еще не писал боту, но ссылка не сформирована "
                "(проверьте переменную SCHOOL_BOT_USERNAME в .env)."
            )

    await message.answer(
        "✅ Ученик добавлен.\n\n"
        f"ФИО: {data['full_name']}\n"
        f"Telegram ID: {data['telegram_id'] if data['telegram_id'] else '-'}\n"
        f"Username: @{data.get('telegram_username') if data.get('telegram_username') else '-'}\n"
        f"Телефон: {phone if phone else '-'}\n"
        f"ID в базе: {student_id}\n"
        f"Роль student: {'создана' if data['telegram_id'] else 'будет создана после входа по ссылке'}"
        f"{referral_text}"
        f"{onboarding_text}",
        reply_markup=get_admin_menu()
    )

    await state.clear()


@router.callback_query(lambda c: c.data == "admin_assign_lesson")
async def admin_assign_lesson(callback: CallbackQuery, state: FSMContext):
    if not is_admin_role(callback.from_user.id):
        await callback.answer("Нет доступа", show_alert=True)
        return

    students = get_all_students()
    if not students:
        await callback.message.answer("Сначала добавь хотя бы одного ученика.")
        await callback.answer()
        return

    text_lines = ["Выбери ученика из списка ниже и отправь его ID (в скобках):\n"]
    for index, student in enumerate(students, start=1):
        student_id, full_name, telegram_id, phone = student
        text_lines.append(f"{index}. {full_name} (ID: {student_id})")

    await callback.message.answer("\n".join(text_lines))
    await state.set_state(AdminStates.choosing_student_for_lesson)
    await callback.answer()


@router.message(AdminStates.choosing_student_for_lesson)
async def choose_student_for_lesson(message: Message, state: FSMContext):
    if not is_admin_role(message.from_user.id):
        await message.answer("Нет доступа.")
        await state.clear()
        return

    text = message.text.strip()

    if not text.isdigit():
        await message.answer("Введи ID ученика числом.")
        return

    student_id = int(text)
    students = get_all_students()
    valid_ids = [student[0] for student in students]

    if student_id not in valid_ids:
        await message.answer("Ученика с таким ID нет. Введи корректный ID.")
        return

    teachers = list_teacher_profiles(limit=1000)
    if not teachers:
        await message.answer("Преподаватели пока не добавлены.")
        await state.clear()
        return

    await state.update_data(
        student_id=student_id,
        assign_teacher_candidates=[int(item[0]) for item in teachers],
    )
    await message.answer(
        "Выберите преподавателя из списка кнопками ниже.\n"
        "Или введите часть ФИО для поиска (например: Ма).",
        reply_markup=get_teacher_selection_keyboard(teachers, action_prefix="assign_teacher_pick"),
    )
    await state.set_state(AdminStates.waiting_teacher_name)
    return


@router.message(AdminStates.waiting_teacher_selection)
async def get_teacher_name(message: Message, state: FSMContext):
    if not is_admin_role(message.from_user.id):
        await message.answer("Нет доступа.")
        await state.clear()
        return

    text = (message.text or "").strip()
    if not text.isdigit():
        await message.answer("Введите ID преподавателя числом из списка выше.")
        return

    teacher_id = int(text)
    data = await state.get_data()
    allowed_ids = data.get("assign_teacher_candidates") or []
    if teacher_id not in allowed_ids:
        await message.answer("Такого ID нет в текущем списке. Введите ID из списка выше.")
        return

    teacher = get_teacher_profile_by_id(teacher_id)
    if not teacher:
        await message.answer("Преподаватель не найден. Повторите поиск.")
        await state.set_state(AdminStates.waiting_teacher_name)
        return

    _, _teacher_telegram_id, teacher_name, _teacher_subject, _description, _photo, _username = teacher
    await state.update_data(teacher_id=teacher_id, teacher_name=teacher_name)
    await message.answer(f"Выбран преподаватель: {teacher_name}\n\nВведите предмет:")
    await state.set_state(AdminStates.waiting_subject_name)
    return


@router.message(AdminStates.waiting_teacher_name)
async def search_teacher_for_lesson_by_fio(message: Message, state: FSMContext):
    if not is_admin_role(message.from_user.id):
        await message.answer("Нет доступа.")
        await state.clear()
        return

    query = (message.text or "").strip().lower()
    if len(query) < 2:
        await message.answer("Введите минимум 2 символа ФИО преподавателя для поиска.")
        return

    teachers = search_teacher_profiles(query, limit=50)
    teachers = [item for item in teachers if query in (item[1] or "").lower()]
    if not teachers:
        await message.answer("По ФИО преподаватели не найдены. Попробуйте другой запрос.")
        return

    await state.update_data(assign_teacher_candidates=[int(item[0]) for item in teachers])
    await message.answer(
        "Найдены преподаватели. Выберите нужного кнопкой:",
        reply_markup=get_teacher_selection_keyboard(teachers, action_prefix="assign_teacher_pick"),
    )
    await state.set_state(AdminStates.waiting_teacher_name)


@router.callback_query(lambda c: c.data.startswith("assign_teacher_pick_"))
async def choose_teacher_for_lesson(callback: CallbackQuery, state: FSMContext):
    if not is_admin_role(callback.from_user.id):
        await callback.answer("Нет доступа", show_alert=True)
        await state.clear()
        return

    try:
        teacher_id = int(callback.data.split("_")[-1])
    except (TypeError, ValueError):
        await callback.answer("Не удалось определить преподавателя", show_alert=True)
        return

    data = await state.get_data()
    allowed_ids = data.get("assign_teacher_candidates") or []
    if allowed_ids and teacher_id not in allowed_ids:
        await callback.answer("Преподаватель не из текущего списка", show_alert=True)
        return

    teacher = get_teacher_profile_by_id(teacher_id)
    if not teacher:
        await callback.answer("Преподаватель не найден", show_alert=True)
        return

    _, _teacher_telegram_id, teacher_name, _teacher_subject, _description, _photo, _username = teacher
    await state.update_data(teacher_id=teacher_id, teacher_name=teacher_name)
    subjects = [item for item in get_teacher_catalog_subjects() if item]
    await state.update_data(
        assign_subject_options=subjects,
        assign_subject_waiting_new=False,
        assign_subject_waiting_alias=False,
        assign_subject_base=None,
    )
    if subjects:
        await callback.message.answer(
            f"Выбран преподаватель: {teacher_name}\n\n"
            "Выберите предмет кнопкой из списка ниже\n"
            "или введите часть названия для фильтрации.",
            reply_markup=get_subject_selection_keyboard(subjects),
        )
    else:
        await callback.message.answer(
            f"Выбран преподаватель: {teacher_name}\n\n"
            "Справочник предметов пока пуст.\n"
            "Введите новый предмет текстом:"
        )
        await state.update_data(assign_subject_waiting_new=True)
    await state.set_state(AdminStates.waiting_subject_name)
    await callback.answer()


@router.callback_query(
    AdminStates.waiting_subject_name,
    lambda c: c.data.startswith("assign_subject_pick_") or c.data == "assign_subject_add_new",
)
async def process_assign_subject_pick(callback: CallbackQuery, state: FSMContext):
    if not is_admin_role(callback.from_user.id):
        await callback.answer("Нет доступа", show_alert=True)
        await state.clear()
        return

    data = await state.get_data()
    subject_options = data.get("assign_subject_options") or []

    if callback.data == "assign_subject_add_new":
        await state.update_data(
            assign_subject_waiting_new=True,
            assign_subject_waiting_alias=False,
            assign_subject_base=None,
        )
        await callback.message.answer("Введите новый предмет текстом:")
        await callback.answer()
        return

    try:
        subject_index = int(callback.data.split("_")[-1])
    except (TypeError, ValueError):
        await callback.answer("Не удалось определить предмет", show_alert=True)
        return

    if subject_index < 0 or subject_index >= len(subject_options):
        await callback.answer("Предмет не найден в текущем списке", show_alert=True)
        return

    selected_subject = (subject_options[subject_index] or "").strip()
    if not selected_subject:
        await callback.answer("Некорректный предмет", show_alert=True)
        return

    await state.update_data(
        assign_subject_base=selected_subject,
        assign_subject_waiting_new=False,
        assign_subject_waiting_alias=False,
    )
    await callback.message.answer(
        f"Выбран предмет: {selected_subject}\n\n"
        "Нужно переименовать его только для этого ученика?",
        reply_markup=get_assign_subject_rename_keyboard(),
    )
    await callback.answer()


@router.callback_query(
    AdminStates.waiting_subject_name,
    lambda c: c.data in {"assign_subject_keep", "assign_subject_rename"},
)
async def process_assign_subject_rename_choice(callback: CallbackQuery, state: FSMContext):
    if not is_admin_role(callback.from_user.id):
        await callback.answer("Нет доступа", show_alert=True)
        await state.clear()
        return

    data = await state.get_data()
    base_subject = (data.get("assign_subject_base") or "").strip()
    if not base_subject:
        await callback.answer("Сначала выберите предмет", show_alert=True)
        return

    if callback.data == "assign_subject_keep":
        await state.update_data(
            subject_name=base_subject,
            expect_custom_subject=False,
            assign_subject_waiting_alias=False,
            assign_subject_waiting_new=False,
        )
        await callback.message.answer("Выбери тип тарифа:", reply_markup=get_tariff_keyboard())
        await state.set_state(AdminStates.waiting_tariff_type)
        await callback.answer()
        return

    await state.update_data(
        assign_subject_waiting_alias=True,
        assign_subject_waiting_new=False,
    )
    await callback.message.answer(
        f"Введите новое отображаемое название для ученика.\n"
        f"Базовый предмет останется: {base_subject}\n\n"
        "Пример: ИЗО (на холсте)"
    )
    await callback.answer()


@router.message(AdminStates.waiting_subject_name)
async def get_subject_name(message: Message, state: FSMContext):
    if not is_admin_role(message.from_user.id):
        await message.answer("Нет доступа.")
        await state.clear()
        return

    raw_text = (message.text or "").strip()
    if len(raw_text) < 2:
        await message.answer("Введите минимум 2 символа.")
        return

    data = await state.get_data()
    waiting_new = bool(data.get("assign_subject_waiting_new"))
    waiting_alias = bool(data.get("assign_subject_waiting_alias"))

    if waiting_alias:
        await state.update_data(
            subject_name=raw_text,
            expect_custom_subject=True,
            assign_subject_waiting_alias=False,
        )
        await message.answer("Выбери тип тарифа:", reply_markup=get_tariff_keyboard())
        await state.set_state(AdminStates.waiting_tariff_type)
        return

    if waiting_new:
        await state.update_data(
            subject_name=raw_text,
            expect_custom_subject=True,
            assign_subject_waiting_new=False,
        )
        await message.answer("Выбери тип тарифа:", reply_markup=get_tariff_keyboard())
        await state.set_state(AdminStates.waiting_tariff_type)
        return

    subject_query = raw_text.lower()
    subjects = [item for item in get_teacher_catalog_subjects() if item]
    matched_subjects = [item for item in subjects if subject_query in item.lower()]
    await state.update_data(assign_subject_options=matched_subjects)

    if matched_subjects:
        await message.answer(
            "Найдены предметы. Выберите кнопкой или добавьте новый:",
            reply_markup=get_subject_selection_keyboard(matched_subjects),
        )
        return

    await message.answer(
        "По запросу предметы не найдены.\n"
        "Введите другой запрос или нажмите «Добавить новый предмет».",
        reply_markup=get_subject_selection_keyboard([]),
    )


@router.callback_query(AdminStates.waiting_tariff_type, lambda c: c.data in ["tariff_single", "tariff_package"])
async def choose_tariff_type(callback: CallbackQuery, state: FSMContext):
    if not is_admin_role(callback.from_user.id):
        await callback.answer("Нет доступа", show_alert=True)
        await state.clear()
        return

    tariff_map = {
        "tariff_single": "single",
        "tariff_package": "package"
    }

    tariff_type = tariff_map[callback.data]
    await state.update_data(tariff_type=tariff_type)

    await callback.message.answer("Сколько занятий начислить на баланс?")
    await state.set_state(AdminStates.waiting_lesson_balance)
    await callback.answer()


@router.message(AdminStates.waiting_lesson_balance)
async def get_lesson_balance(message: Message, state: FSMContext):
    if not is_admin_role(message.from_user.id):
        await message.answer("Нет доступа.")
        await state.clear()
        return

    text = message.text.strip()

    if not text.isdigit():
        await message.answer("Введи количество занятий числом.")
        return

    lesson_balance = int(text)
    if lesson_balance < 0:
        await message.answer("Количество занятий не может быть отрицательным.")
        return

    data = await state.get_data()

    teacher_id = data.get("teacher_id")
    if not teacher_id:
        teacher_id = add_teacher_if_not_exists(data["teacher_name"])

    add_student_lesson(
        student_id=data["student_id"],
        teacher_id=teacher_id,
        subject_name=data["subject_name"],
        lesson_balance=lesson_balance,
        tariff_type=data["tariff_type"]
    )
    log_admin_action(
        admin_telegram_id=message.from_user.id,
        action_type="assign_lesson",
        target_type="student_lesson",
        target_id=None,
        details={
            "student_id": data["student_id"],
            "teacher_id": teacher_id,
            "subject_name": data["subject_name"],
            "lesson_balance": lesson_balance,
            "tariff_type": data["tariff_type"],
        },
        status="success",
    )

    await message.answer(
        "✅ Направление добавлено.\n\n"
        f"ID ученика: {data['student_id']}\n"
        f"Преподаватель: {data['teacher_name']}\n"
        f"Предмет: {data['subject_name']}\n"
        f"Тариф: {data['tariff_type']}\n"
        f"Баланс: {lesson_balance}",
        reply_markup=get_admin_menu()
    )

    await state.clear()


@router.callback_query(lambda c: c.data == "admin_find_student")
async def admin_find_student(callback: CallbackQuery, state: FSMContext):
    if not is_admin_role(callback.from_user.id):
        await callback.answer("Нет доступа", show_alert=True)
        return

    await callback.message.answer("Введи имя или часть имени ученика:")
    await state.set_state(AdminStates.waiting_student_search)
    await callback.answer()


@router.message(AdminStates.waiting_student_search)
async def search_student(message: Message, state: FSMContext):
    if not is_admin_role(message.from_user.id):
        await message.answer("Нет доступа.")
        await state.clear()
        return

    search_text = message.text.strip()
    students = find_students_by_name_with_username(search_text)

    if not students:
        await message.answer("Ничего не найдено.", reply_markup=get_admin_menu())
        await state.clear()
        return

    if len(students) > 1:
        await message.answer(
            "Найдено несколько учеников. Выберите нужного:",
            reply_markup=get_student_disambiguation_keyboard(students, action_prefix="find_student_pick"),
        )
        await state.clear()
        return

    result_messages = []

    for student in students:
        student_id, full_name, telegram_id, phone, telegram_username = student
        directions = get_student_directions(student_id)
        username_text = f"@{telegram_username}" if telegram_username else "-"
        await send_student_contact_shortcut(
            message,
            telegram_id=telegram_id,
            username=telegram_username,
        )
        contact_keyboard = None

        text = (
            f"👤 <b>{full_name}</b>\n"
            f"🆔 ID: <code>{student_id}</code>\n"
            f"🔗 Username: <code>{username_text}</code>\n"
            f"📱 Телефон: {phone if phone else '-'}\n"
            f"🔗 Telegram ID: {telegram_id if telegram_id else '-'}\n\n"
        )

        if directions:
            text += "<b>Направления:</b>\n"
            for direction in directions:
                _, teacher_name, subject_name, lesson_balance, tariff_type = direction
                tariff_text = "Разовое" if tariff_type == "single" else "Пакет"
                text += (
                    f"• {subject_name} — {teacher_name}\n"
                    f"  Тариф: {tariff_text}\n"
                    f"  Остаток: {lesson_balance}\n"
                )
        else:
            text += "Направлений пока нет."

        result_messages.append(text)

    for text in result_messages:
        await message.answer(text, parse_mode="HTML", reply_markup=contact_keyboard)

    await message.answer("Поиск завершен.", reply_markup=get_admin_menu())
    await state.clear()


@router.callback_query(lambda c: c.data.startswith("find_student_pick_"))
async def find_student_pick_from_disambiguation(callback: CallbackQuery):
    if not is_admin_role(callback.from_user.id):
        await callback.answer("Нет доступа", show_alert=True)
        return

    try:
        student_id = int(callback.data.split("_")[-1])
    except (TypeError, ValueError):
        await callback.answer("Не удалось определить ученика", show_alert=True)
        return

    student = get_student_by_id_with_username(student_id)
    if not student:
        await callback.answer("Ученик не найден", show_alert=True)
        return

    _student_id, full_name, telegram_id, phone, telegram_username = student
    directions = get_student_directions(student_id)
    username_text = f"@{telegram_username}" if telegram_username else "-"
    await send_student_contact_shortcut(
        callback.message,
        telegram_id=telegram_id,
        username=telegram_username,
    )
    contact_keyboard = None

    text = (
        f"👤 <b>{full_name}</b>\n"
        f"🆔 ID: <code>{student_id}</code>\n"
        f"🔗 Username: <code>{username_text}</code>\n"
        f"📱 Телефон: {phone if phone else '-'}\n"
        f"🔗 Telegram ID: {telegram_id if telegram_id else '-'}\n\n"
    )

    if directions:
        text += "<b>Направления:</b>\n"
        for direction in directions:
            _, teacher_name, subject_name, lesson_balance, tariff_type = direction
            tariff_text = "Разовое" if tariff_type == "single" else "Пакет"
            text += (
                f"• {subject_name} — {teacher_name}\n"
                f"  Тариф: {tariff_text}\n"
                f"  Остаток: {lesson_balance}\n"
            )
    else:
        text += "Направлений пока нет."

    await callback.message.answer(text, parse_mode="HTML", reply_markup=contact_keyboard)
    await callback.message.answer("Поиск завершен.", reply_markup=get_admin_reply_menu(callback.from_user.id))
    await callback.answer()


@router.callback_query(lambda c: c.data == "admin_attendance")
async def admin_attendance(callback: CallbackQuery, state: FSMContext):
    if not is_admin_role(callback.from_user.id):
        await callback.answer("Нет доступа", show_alert=True)
        return

    await update_flow_message(
        callback,
        "Введите имя или часть имени ученика для отметки посещения.\n\n"
        "Финальный результат будет отправлен отдельным сообщением.",
    )
    await state.set_state(AdminStates.waiting_attendance_student_search)
    await callback.answer()


@router.callback_query(lambda c: c.data == "teacher_students")
async def teacher_students(callback: CallbackQuery):
    if not is_teacher_role(callback.from_user.id):
        await callback.answer("Нет доступа", show_alert=True)
        return

    teacher = get_teacher_by_telegram_id(callback.from_user.id)
    if not teacher:
        await callback.message.answer(
            "Профиль преподавателя не найден. Пожалуйста, обратитесь к администратору."
        )
        await callback.answer()
        return

    students = get_students_by_teacher_telegram_id(callback.from_user.id)
    if not students:
        await callback.message.answer("За вами пока не закреплены ученики.")
        await callback.answer()
        return

    lines = ["<b>Ваши ученики:</b>\n"]
    for student_id, full_name, _telegram_id, phone, telegram_username in students:
        directions = get_teacher_owned_directions(callback.from_user.id, student_id)
        direction_text = "; ".join(
            f"{subject_name} (остаток: {lesson_balance})"
            for _, _, subject_name, lesson_balance, _ in directions
        ) or "Направления пока не найдены"
        username_text = f"@{telegram_username}" if telegram_username else "не указан"

        lines.append(
            f"• <b>{full_name}</b>\n"
            f"Username: <code>{username_text}</code>\n"
            f"Телефон: {phone if phone else '-'}\n"
            f"Направления: {direction_text}\n"
        )

    await callback.message.answer("\n".join(lines), parse_mode="HTML", reply_markup=get_teacher_menu())
    await callback.answer()


@router.callback_query(lambda c: c.data == "teacher_weekly_report")
async def teacher_weekly_report(callback: CallbackQuery):
    if not is_teacher_role(callback.from_user.id):
        await callback.answer("Нет доступа", show_alert=True)
        return

    rows = get_weekly_lessons_report_for_teacher_telegram(callback.from_user.id)
    if not rows:
        await callback.message.answer(
            "📊 <b>Отчёт за неделю</b>\n\nЗа последние 7 дней проведённых занятий не найдено.",
            parse_mode="HTML",
            reply_markup=get_teacher_menu(),
        )
        await callback.answer()
        return

    total = sum(int(r[3] or 0) for r in rows)
    lines = [
        "📊 <b>Ваш отчёт за последние 7 дней</b>",
        "",
        f"Всего проведено занятий: <b>{total}</b>",
        "",
        "<b>По предметам:</b>",
    ]
    for _, _, subject_name, lessons_count, last_lesson_date in rows:
        last_view = (last_lesson_date or "—")[:10] if last_lesson_date else "—"
        lines.append(
            f"   • {subject_name}: <b>{int(lessons_count or 0)}</b>"
            f"  (последнее: {last_view})"
        )

    await callback.message.answer(
        "\n".join(lines),
        parse_mode="HTML",
        reply_markup=get_teacher_menu(),
    )
    await callback.answer()


@router.callback_query(lambda c: c.data == "teacher_attendance")
async def teacher_attendance_v2(callback: CallbackQuery):
    if not is_teacher_role(callback.from_user.id):
        await callback.answer("Нет доступа", show_alert=True)
        return

    students = get_students_by_teacher_telegram_id(callback.from_user.id)
    if not students:
        await update_flow_message(callback, "За Вами пока не закреплены ученики.")
        await callback.answer()
        return

    await update_flow_message(
        callback,
        "Выберите ученика для отметки посещаемости:",
        reply_markup=get_teacher_attendance_students_keyboard(students),
    )
    await callback.answer()


@router.callback_query(lambda c: c.data.startswith("teacher_attendance_student_"))
async def teacher_attendance_choose_student(callback: CallbackQuery):
    if not is_teacher_role(callback.from_user.id):
        await callback.answer("Нет доступа", show_alert=True)
        return

    try:
        student_id = int(callback.data.split("_")[-1])
    except (TypeError, ValueError):
        await callback.answer("Не удалось определить ученика", show_alert=True)
        return

    directions = get_teacher_owned_directions(callback.from_user.id, student_id)
    if not directions:
        await update_flow_message(callback, "Для этого ученика у Вас пока нет направлений.")
        await callback.answer()
        return

    if len(directions) == 1:
        direction_id = directions[0][0]
        lesson = get_student_lesson_by_id(direction_id)
        if lesson:
            _, _, _, subject_name, lesson_balance, tariff_type, student_name, teacher_name = lesson
            await update_flow_message(
                callback,
                f"Ученик: {student_name}\n"
                f"Предмет: {subject_name}\n"
                f"Преподаватель: {teacher_name}\n"
                f"Остаток: {lesson_balance}\n\n"
                f"Отметьте посещение:",
                reply_markup=get_attendance_mark_keyboard(direction_id),
            )
            await callback.answer()
            return

    await update_flow_message(
        callback,
        "Выберите направление ученика для отметки посещаемости:",
        reply_markup=get_attendance_direction_keyboard(directions),
    )
    await callback.answer()


@router.callback_query(lambda c: c.data == "teacher_attendance_legacy")
async def teacher_attendance(callback: CallbackQuery):
    if not is_teacher_role(callback.from_user.id):
        await callback.answer("Нет доступа", show_alert=True)
        return

    students = get_students_by_teacher_telegram_id(callback.from_user.id)
    if not students:
        await callback.message.answer("За вами пока не закреплены ученики.")
        await callback.answer()
        return

    directions = []
    for student in students:
        student_id = student[0]
        directions.extend(get_teacher_owned_directions(callback.from_user.id, student_id))

    if not directions:
        await callback.message.answer("Для вас пока нет направлений для отметки посещения.")
        await callback.answer()
        return

    await callback.message.answer(
        "Выберите направление для отметки посещения:",
        reply_markup=get_attendance_direction_keyboard(directions)
    )
    await callback.answer()


@router.message(AdminStates.waiting_attendance_student_search)
async def attendance_student_search(message: Message, state: FSMContext):
    if not is_admin_role(message.from_user.id):
        await message.answer("Нет доступа.")
        await state.clear()
        return

    search_text = message.text.strip()
    students = find_students_by_name_with_username(search_text)

    if not students:
        await message.answer("Ученик не найден.", reply_markup=get_admin_menu())
        await state.clear()
        return

    if len(students) > 1:
        await message.answer(
            "Найдено несколько учеников. Выберите нужного:",
            reply_markup=get_student_disambiguation_keyboard(students, action_prefix="attendance_pick_student"),
        )
        await state.clear()
        return

    student_id, full_name, telegram_id, phone, telegram_username = students[0]
    directions = get_student_directions(student_id)

    if not directions:
        await message.answer("У этого ученика пока нет направлений.", reply_markup=get_admin_menu())
        await state.clear()
        return

    await send_student_contact_shortcut(
        message,
        telegram_id=telegram_id,
        username=telegram_username,
    )
    contact_keyboard = None
    if contact_keyboard:
        await message.answer("Быстрый переход в чат с учеником:", reply_markup=contact_keyboard)

    await message.answer(
        f"Выбери направление для ученика {full_name}:",
        reply_markup=get_attendance_direction_keyboard(directions)
    )
    await state.clear()


@router.callback_query(lambda c: c.data.startswith("attendance_pick_student_"))
async def attendance_pick_student_from_disambiguation(callback: CallbackQuery):
    if not is_admin_role(callback.from_user.id):
        await callback.answer("Нет доступа", show_alert=True)
        return

    try:
        student_id = int(callback.data.split("_")[-1])
    except (TypeError, ValueError):
        await callback.answer("Не удалось определить ученика", show_alert=True)
        return

    student = get_student_by_id_with_username(student_id)
    if not student:
        await callback.answer("Ученик не найден", show_alert=True)
        return

    _id, full_name, telegram_id, _phone, telegram_username = student
    directions = get_student_directions(student_id)
    if not directions:
        await callback.message.answer("У этого ученика пока нет направлений.", reply_markup=get_admin_reply_menu(callback.from_user.id))
        await callback.answer()
        return

    await send_student_contact_shortcut(
        callback.message,
        telegram_id=telegram_id,
        username=telegram_username,
    )
    contact_keyboard = None
    if contact_keyboard:
        await callback.message.answer("Быстрый переход в чат с учеником:", reply_markup=contact_keyboard)

    await callback.message.answer(
        f"Выбери направление для ученика {full_name}:",
        reply_markup=get_attendance_direction_keyboard(directions),
    )
    await callback.answer()


@router.callback_query(lambda c: c.data.startswith("attendance_direction_"))
async def choose_attendance_direction(callback: CallbackQuery):
    direction_id = int(callback.data.split("_")[-1])

    if not can_manage_attendance(callback.from_user.id, direction_id):
        await callback.answer("Нет доступа", show_alert=True)
        return

    lesson = get_student_lesson_by_id(direction_id)
    if not lesson:
        await callback.answer("Направление не найдено", show_alert=True)
        return

    _, _, _, subject_name, lesson_balance, tariff_type, student_name, teacher_name = lesson

    await update_flow_message(
        callback,
        f"Ученик: {student_name}\n"
        f"Предмет: {subject_name}\n"
        f"Преподаватель: {teacher_name}\n"
        f"Остаток: {lesson_balance}\n\n"
        f"Отметьте посещение:",
        reply_markup=get_attendance_mark_keyboard(direction_id)
    )
    await callback.answer()


@router.callback_query(lambda c: c.data.startswith("attendance_present_") or c.data.startswith("attendance_absent_"))
async def mark_student_attendance(callback: CallbackQuery):
    if callback.data.startswith("attendance_present_"):
        direction_id = int(callback.data.split("_")[-1])
        status = "present"
    else:
        direction_id = int(callback.data.split("_")[-1])
        status = "absent"

    if not can_manage_attendance(callback.from_user.id, direction_id):
        await callback.answer("Нет доступа", show_alert=True)
        return

    lesson = get_student_lesson_by_id(direction_id)
    if not lesson:
        await callback.answer("Направление не найдено", show_alert=True)
        return

    _, student_id, _, subject_name, lesson_balance_before, tariff_type, student_name, teacher_name = lesson

    if status == "present" and has_recent_attendance(direction_id, within_minutes=5):
        confirm_kb = InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(
                text="✅ Да, списать повторно",
                callback_data=f"attendance_confirm_present_{direction_id}",
            )],
            [InlineKeyboardButton(
                text="❌ Отмена",
                callback_data=f"attendance_mark_cancel_{direction_id}",
            )],
        ])
        await callback.message.edit_text(
            f"⚠️ <b>Занятие уже отмечено менее 5 минут назад!</b>\n\n"
            f"Ученик: <b>{student_name}</b> — {subject_name}\n\n"
            f"Списать повторно?",
            parse_mode="HTML",
            reply_markup=confirm_kb,
        )
        await callback.answer()
        return

    await callback.answer()

    lesson_datetime = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    attendance_id = mark_attendance(direction_id, status, callback.from_user.id)
    log_admin_action(
        admin_telegram_id=callback.from_user.id,
        action_type="mark_attendance",
        target_type="student_lesson",
        target_id=direction_id,
        details={"status": status},
        status="success",
    )

    updated_lesson = get_student_lesson_by_id(direction_id)
    _, _, _, _, lesson_balance_after, _, _, _ = updated_lesson

    marked_by_name = callback.from_user.full_name or str(callback.from_user.id)
    asyncio.create_task(_sync_attendance_to_sheets(
        attendance_id=attendance_id,
        lesson_datetime=lesson_datetime,
        teacher_name=teacher_name,
        student_name=student_name,
        subject_name=subject_name,
        tariff_type=tariff_type,
        status=status,
        balance_before=lesson_balance_before,
        balance_after=lesson_balance_after,
        marked_by_name=marked_by_name,
    ))
    asyncio.create_task(_update_summary_sheets_bg())

    student = get_student_by_id(student_id)
    student_telegram_id = student[2] if student else None
    student_max_id = get_student_max_id(student_id)

    await notify_student_about_attendance_clean(
        callback,
        student_telegram_id=student_telegram_id,
        student_max_id=student_max_id,
        student_name=student_name,
        subject_name=subject_name,
        teacher_name=teacher_name,
        tariff_type=tariff_type,
        status=status,
        lesson_balance_before=lesson_balance_before,
        lesson_balance_after=lesson_balance_after,
    )

    teacher_telegram_id = None
    teacher = get_teacher_by_id(lesson[2])
    if teacher:
        teacher_telegram_id = teacher[1]
    if teacher_telegram_id and teacher_telegram_id != callback.from_user.id:
        await notify_teacher_about_attendance(
            callback,
            teacher_telegram_id=teacher_telegram_id,
            student_name=student_name,
            subject_name=subject_name,
            status=status,
            lesson_balance_after=lesson_balance_after,
        )

    status_text = "Был" if status == "present" else "Не был"
    status_emoji = "✅" if status == "present" else "❌"

    confirmation_text = (
        f"{status_emoji} <b>Посещаемость отмечена</b>\n\n"
        f"👤 <b>Ученик:</b> {student_name}\n"
        f"📚 <b>Предмет:</b> {subject_name}\n"
        f"👨‍🏫 <b>Преподаватель:</b> {teacher_name}\n"
        f"📊 <b>Статус:</b> {status_text}\n"
        f"💾 <b>Баланс был:</b> {lesson_balance_before}\n"
        f"💾 <b>Баланс стал:</b> {lesson_balance_after}\n\n"
        f"<i>Отметка сохранена в истории ✓</i>"
    )

    try:
        await callback.message.edit_text(confirmation_text, parse_mode="HTML", reply_markup=None)
    except Exception:
        await callback.message.answer(confirmation_text, parse_mode="HTML")

    reply_kb = get_admin_reply_menu(callback.from_user.id) if is_admin_role(callback.from_user.id) else get_teacher_menu()
    await callback.message.answer("📋 Меню:", reply_markup=reply_kb)


@router.callback_query(lambda c: c.data.startswith("attendance_confirm_present_"))
async def attendance_confirm_present(callback: CallbackQuery):
    direction_id = int(callback.data.split("_")[-1])
    if not can_manage_attendance(callback.from_user.id, direction_id):
        await callback.answer("Нет доступа", show_alert=True)
        return
    lesson = get_student_lesson_by_id(direction_id)
    if not lesson:
        await callback.answer("Направление не найдено", show_alert=True)
        return
    _, student_id, _, subject_name, lesson_balance_before, tariff_type, student_name, teacher_name = lesson
    lesson_datetime = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    attendance_id = mark_attendance(direction_id, "present", callback.from_user.id)
    log_admin_action(
        admin_telegram_id=callback.from_user.id,
        action_type="mark_attendance",
        target_type="student_lesson",
        target_id=direction_id,
        details={"status": "present", "forced_repeat": True},
        status="success",
    )
    updated_lesson = get_student_lesson_by_id(direction_id)
    _, _, _, _, lesson_balance_after, _, _, _ = updated_lesson
    marked_by_name = callback.from_user.full_name or str(callback.from_user.id)
    asyncio.create_task(_sync_attendance_to_sheets(
        attendance_id=attendance_id,
        lesson_datetime=lesson_datetime,
        teacher_name=teacher_name,
        student_name=student_name,
        subject_name=subject_name,
        tariff_type=tariff_type,
        status="present",
        balance_before=lesson_balance_before,
        balance_after=lesson_balance_after,
        marked_by_name=marked_by_name,
    ))
    asyncio.create_task(_update_summary_sheets_bg())
    student = get_student_by_id(student_id)
    student_telegram_id = student[2] if student else None
    student_max_id = get_student_max_id(student_id)
    await notify_student_about_attendance_clean(
        callback,
        student_telegram_id=student_telegram_id,
        student_max_id=student_max_id,
        student_name=student_name,
        subject_name=subject_name,
        teacher_name=teacher_name,
        tariff_type=tariff_type,
        status="present",
        lesson_balance_before=lesson_balance_before,
        lesson_balance_after=lesson_balance_after,
    )
    confirmation_text = (
        f"✅ <b>Повторное списание выполнено</b>\n\n"
        f"👤 <b>Ученик:</b> {student_name}\n"
        f"📚 <b>Предмет:</b> {subject_name}\n"
        f"💾 <b>Баланс был:</b> {lesson_balance_before} → <b>стал:</b> {lesson_balance_after}"
    )
    try:
        await callback.message.edit_text(confirmation_text, parse_mode="HTML", reply_markup=None)
    except Exception:
        await callback.message.answer(confirmation_text, parse_mode="HTML")
    reply_kb = get_admin_reply_menu(callback.from_user.id) if is_admin_role(callback.from_user.id) else get_teacher_menu()
    await callback.message.answer("📋 Меню:", reply_markup=reply_kb)
    await callback.answer()


@router.callback_query(lambda c: c.data.startswith("attendance_mark_cancel_"))
async def attendance_mark_cancel(callback: CallbackQuery):
    direction_id = int(callback.data.split("_")[-1])
    lesson = get_student_lesson_by_id(direction_id)
    if not lesson:
        await callback.answer("Направление не найдено", show_alert=True)
        return
    _, student_id, _, subject_name, lesson_balance, tariff_type, student_name, teacher_name = lesson
    try:
        await callback.message.edit_text(
            f"Ученик: {student_name}\n"
            f"Предмет: {subject_name}\n"
            f"Преподаватель: {teacher_name}\n"
            f"Остаток: {lesson_balance}\n\n"
            f"Отметьте посещение:",
            reply_markup=get_attendance_mark_keyboard(direction_id),
        )
    except Exception:
        pass
    await callback.answer("Отменено")


@router.callback_query(lambda c: c.data == "admin_add_balance")
async def admin_add_balance(callback: CallbackQuery, state: FSMContext):
    if not is_admin_role(callback.from_user.id):
        await callback.answer("Нет доступа", show_alert=True)
        return

    await callback.message.answer("Введи имя или часть имени ученика для корректировки баланса:")
    await state.set_state(AdminStates.waiting_balance_student_search)
    await callback.answer()


@router.message(AdminStates.waiting_balance_student_search)
async def balance_student_search(message: Message, state: FSMContext):
    if not is_admin_role(message.from_user.id):
        await message.answer("Нет доступа.")
        await state.clear()
        return

    search_text = message.text.strip()
    students = find_students_by_name_with_username(search_text)

    if not students:
        await message.answer("Ученик не найден.", reply_markup=get_admin_menu())
        await state.clear()
        return

    if len(students) > 1:
        await message.answer(
            "Найдено несколько учеников. Выберите нужного:",
            reply_markup=get_student_disambiguation_keyboard(students, action_prefix="balance_pick_student"),
        )
        await state.clear()
        return

    student_id, full_name, telegram_id, phone, telegram_username = students[0]
    directions = get_student_directions(student_id)

    if not directions:
        await message.answer("У этого ученика пока нет направлений.", reply_markup=get_admin_menu())
        await state.clear()
        return

    await send_student_contact_shortcut(
        message,
        telegram_id=telegram_id,
        username=telegram_username,
    )
    contact_keyboard = None
    if contact_keyboard:
        await message.answer("Быстрый переход в чат с учеником:", reply_markup=contact_keyboard)

    await message.answer(
        f"Выбери направление для корректировки баланса ученика {full_name}:",
        reply_markup=get_balance_direction_keyboard(directions)
    )
    await state.clear()


@router.callback_query(lambda c: c.data.startswith("balance_pick_student_"))
async def balance_pick_student_from_disambiguation(callback: CallbackQuery):
    try:
        if not is_admin_role(callback.from_user.id):
            await callback.answer("Нет доступа", show_alert=True)
            return

        try:
            student_id = int(callback.data.split("_")[-1])
        except (TypeError, ValueError):
            await callback.answer("Не удалось определить ученика", show_alert=True)
            return

        student = get_student_by_id_with_username(student_id)
        if student:
            _id, full_name, telegram_id, _phone, telegram_username = student
        else:
            fallback_student = get_student_by_id(student_id)
            if not fallback_student:
                await callback.answer("Ученик не найден", show_alert=True)
                return
            _id, full_name, telegram_id, _phone = fallback_student
            telegram_username = None

        directions = get_student_directions(student_id)
        if not directions:
            await callback.message.answer(
                "У этого ученика пока нет направлений.",
                reply_markup=get_admin_reply_menu(callback.from_user.id),
            )
            await callback.answer()
            return

        await send_student_contact_shortcut(
            callback.message,
            telegram_id=telegram_id,
            username=telegram_username,
        )
        contact_keyboard = None
        if contact_keyboard:
            await callback.message.answer("Быстрый переход в чат с учеником:", reply_markup=contact_keyboard)

        await callback.message.answer(
            f"Выбери направление для корректировки баланса ученика {full_name}:",
            reply_markup=get_balance_direction_keyboard(directions),
        )
        await callback.answer()
    except Exception as exc:
        logger.exception("balance_pick_student_from_disambiguation failed: %s", exc)
        try:
            await callback.answer("Ошибка при выборе ученика", show_alert=True)
        except Exception:
            pass
        await callback.message.answer(
            "Не удалось открыть направления для корректировки баланса. Попробуйте ещё раз или отправьте другой запрос.",
            reply_markup=get_admin_reply_menu(callback.from_user.id) if is_admin_role(callback.from_user.id) else get_main_menu_shortcut_keyboard(),
        )


@router.callback_query(lambda c: c.data.startswith("balance_direction_"))
async def choose_balance_direction(callback: CallbackQuery):
    if not is_admin_role(callback.from_user.id):
        await callback.answer("Нет доступа", show_alert=True)
        return

    direction_id = int(callback.data.split("_")[-1])

    lesson = get_student_lesson_by_id(direction_id)
    if not lesson:
        await callback.answer("Направление не найдено", show_alert=True)
        return

    _, _, _, subject_name, lesson_balance, tariff_type, student_name, teacher_name = lesson

    await callback.message.answer(
        f"Ученик: {student_name}\n"
        f"Предмет: {subject_name}\n"
        f"Преподаватель: {teacher_name}\n"
        f"Текущий баланс: {lesson_balance}\n\n"
        f"Выберите, на сколько изменить баланс (+ или -):",
        reply_markup=get_balance_add_keyboard(direction_id)
    )
    await callback.answer()


@router.callback_query(lambda c: c.data.startswith("balance_add_"))
async def add_balance_to_direction(callback: CallbackQuery):
    if not is_admin_role(callback.from_user.id):
        await callback.answer("Нет доступа", show_alert=True)
        return

    parts = callback.data.split("_")
    direction_id = int(parts[2])
    lessons_to_add = int(parts[3])

    lesson = get_student_lesson_by_id(direction_id)
    if not lesson:
        await callback.answer("Направление не найдено", show_alert=True)
        return

    _, student_id, _, subject_name, lesson_balance_before, tariff_type, student_name, teacher_name = lesson

    add_lessons_to_balance(
        direction_id,
        lessons_to_add,
        created_by=callback.from_user.id,
        comment="Ручная корректировка баланса админом"
    )
    log_admin_action(
        admin_telegram_id=callback.from_user.id,
        action_type="manual_balance_adjust",
        target_type="student_lesson",
        target_id=direction_id,
        details={"lessons_delta": lessons_to_add},
        status="success",
    )

    updated_lesson = get_student_lesson_by_id(direction_id)
    _, _, _, _, lesson_balance_after, _, _, _ = updated_lesson

    operation_text = "Начислено" if lessons_to_add > 0 else "Убавлено"
    operation_amount = abs(lessons_to_add)

    notify_kb = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="🔔 Оповестить ученика", callback_data=f"bal_notify_{direction_id}")],
        [InlineKeyboardButton(text="🔕 Не уведомлять", callback_data="bal_notify_skip")],
    ])

    await callback.message.answer(
        f"✅ Баланс обновлен\n\n"
        f"Ученик: {student_name}\n"
        f"Предмет: {subject_name}\n"
        f"Преподаватель: {teacher_name}\n"
        f"Баланс был: {lesson_balance_before}\n"
        f"{operation_text}: {operation_amount}\n"
        f"Баланс стал: {lesson_balance_after}",
        reply_markup=notify_kb,
    )

    await callback.answer()


@router.callback_query(lambda c: c.data == "bal_notify_skip")
async def bal_notify_skip(callback: CallbackQuery):
    await callback.message.edit_reply_markup(reply_markup=None)
    await callback.answer("Уведомление не отправлено")


@router.callback_query(lambda c: c.data.startswith("bal_notify_"))
async def bal_notify_send(callback: CallbackQuery):
    if not is_admin_role(callback.from_user.id):
        await callback.answer("Нет доступа", show_alert=True)
        return

    direction_id = int(callback.data.split("bal_notify_")[1])
    lesson = get_student_lesson_by_id(direction_id)
    if not lesson:
        await callback.answer("Направление не найдено", show_alert=True)
        return

    _, student_id, _, subject_name, lesson_balance, _, student_name, teacher_name = lesson
    student = get_student_by_id(student_id)
    student_telegram_id = student[2] if student else None
    student_max_id = get_student_max_id(student_id)

    lines = [
        "Здравствуйте!",
        "",
        f"Ученик: {student_name}",
        f"Предмет: {subject_name}",
        f"Преподаватель: {teacher_name}",
        f"Баланс стал: {lesson_balance}",
    ]

    reply_markup = None
    max_kb = None
    if lesson_balance < 0:
        lines.extend([
            "",
            "❗❗❗🔴 ВНИМАНИЕ! У ВАС ЗАДОЛЖЕННОСТЬ! 🔴❗❗❗",
            f"Размер задолженности: {abs(lesson_balance)} занят.",
            "❗❗❗ Пожалуйста, внесите оплату. ❗❗❗",
        ])
        reply_markup = build_payment_prompt_keyboard_clean()
        from shared.max_api import btn as _mbtn, keyboard as _mkb
        max_kb = _mkb(
            [_mbtn("👤 Личный кабинет", "menu_cabinet")],
            [_mbtn("💸 Погасить долг", "menu_paid")],
        )
    else:
        from shared.max_api import btn as _mbtn, keyboard as _mkb
        max_kb = _mkb([_mbtn("👤 Личный кабинет", "menu_cabinet")])

    text = "\n".join(lines)

    sent = False
    if student_telegram_id:
        await send_student_notification(callback, student_telegram_id, text, reply_markup=reply_markup)
        sent = True
    if student_max_id:
        await _send_max_notification(student_max_id, text, max_kb)
        sent = True

    await callback.message.edit_reply_markup(reply_markup=None)
    await callback.answer("✅ Уведомление отправлено" if sent else "⚠️ Не удалось найти контакт ученика", show_alert=not sent)


@router.callback_query(lambda c: c.data == "admin_balance_history")
async def admin_balance_history(callback: CallbackQuery, state: FSMContext):
    if not is_admin_role(callback.from_user.id):
        await callback.answer("Нет доступа", show_alert=True)
        return

    await callback.message.answer("Введи имя или часть имени ученика для просмотра истории:")
    await state.set_state(AdminStates.waiting_history_student_search)
    await callback.answer()


@router.callback_query(lambda c: c.data == "admin_actions_recent")
async def admin_actions_recent(callback: CallbackQuery):
    if not is_admin_role(callback.from_user.id):
        await callback.answer("Нет доступа", show_alert=True)
        return

    rows = get_recent_admin_actions(30)
    text = format_admin_action_log(rows)

    chunks = []
    current = ""
    for line in text.split("\n"):
        if len(current) + len(line) + 1 > 3500:
            if current:
                chunks.append(current)
            current = line
        else:
            current += ("\n" if current else "") + line

    if current:
        chunks.append(current)

    for chunk in chunks:
        await callback.message.answer(chunk, parse_mode="HTML")

    await callback.answer()


@router.callback_query(lambda c: c.data == "admin_teacher_lessons_report")
async def admin_teacher_lessons_report(callback: CallbackQuery):
    if not is_admin_role(callback.from_user.id):
        await callback.answer("Нет доступа", show_alert=True)
        return
    await callback.message.answer(
        "📊 <b>Отчёт по занятиям преподавателей</b>\n\nВыберите период:",
        parse_mode="HTML",
        reply_markup=get_lessons_report_period_keyboard(back_callback="superadmin_section_reports"),
    )
    await callback.answer()


def _period_dates(period_key: str) -> tuple[str, str, str]:
    """Return (start_date, end_date, label) for a period key."""
    today = datetime.now().date()
    if period_key == "today":
        return today.isoformat(), today.isoformat(), "сегодня"
    if period_key == "week":
        return (today - timedelta(days=7)).isoformat(), today.isoformat(), "за 7 дней"
    if period_key == "month":
        return (today - timedelta(days=30)).isoformat(), today.isoformat(), "за 30 дней"
    if period_key == "curmonth":
        start = today.replace(day=1)
        return start.isoformat(), today.isoformat(), f"за {today.strftime('%B %Y')}"
    return (today - timedelta(days=7)).isoformat(), today.isoformat(), "за 7 дней"


def _format_report_text(rows: list, period_label: str, teacher_id: int | None) -> str:
    if not rows:
        return f"📊 За выбранный период ({period_label}) проведённых занятий не найдено."

    lines = [f"📊 <b>Отчёт по занятиям {period_label}</b>\n"]
    current_teacher = None
    total_all = 0

    for row in rows:
        tid, tname, subject, count, last_date = row[0], row[1], row[2], int(row[3] or 0), row[4]
        student_name = row[5] if len(row) > 5 else None
        total_all += count
        last_view = (last_date or "—")[:10] if last_date else "—"

        if current_teacher != tname:
            if current_teacher is not None:
                lines.append("")
            current_teacher = tname
            if teacher_id is None:
                lines.append(f"<b>👨‍🏫 {tname}</b>")

        if student_name:
            lines.append(f"  • {subject} — {student_name}: <b>{count}</b> зан. (посл.: {last_view})")
        else:
            lines.append(f"  • {subject}: <b>{count}</b> зан. (посл.: {last_view})")

    lines.append(f"\n<b>Итого занятий: {total_all}</b>")
    return "\n".join(lines)


def _make_excel(rows: list, period_label: str) -> bytes:
    from io import BytesIO
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Занятия"

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    headers = ["Преподаватель", "Предмет", "Ученик", "Занятий", "Последнее занятие"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for r_idx, row in enumerate(rows, 2):
        tid, tname, subject, count, last_date = row[0], row[1], row[2], int(row[3] or 0), row[4]
        student_name = row[5] if len(row) > 5 else "—"
        last_view = (last_date or "")[:10] if last_date else "—"
        ws.append([tname, subject, student_name or "—", count, last_view])

    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


@router.callback_query(lambda c: c.data.startswith("lreport_period_") and c.data != "lreport_period_custom")
async def lreport_choose_period(callback: CallbackQuery):
    if not is_admin_role(callback.from_user.id):
        await callback.answer("Нет доступа", show_alert=True)
        return
    period_key = callback.data.removeprefix("lreport_period_")
    teachers = get_teachers_with_lessons()
    if not teachers:
        await callback.message.answer("Нет данных о занятиях.", parse_mode="HTML")
        await callback.answer()
        return
    await callback.message.answer(
        "Выберите преподавателя:",
        reply_markup=get_lessons_report_teacher_filter_keyboard(
            teachers, period_key, back_callback="admin_teacher_lessons_report"
        ),
    )
    await callback.answer()


@router.callback_query(lambda c: c.data == "lreport_period_custom")
async def lreport_period_custom(callback: CallbackQuery, state: FSMContext):
    if not is_admin_role(callback.from_user.id):
        await callback.answer("Нет доступа", show_alert=True)
        return
    await callback.message.answer(
        "Введите <b>дату начала</b> периода в формате ДД.ММ.ГГГГ или ГГГГ-ММ-ДД:",
        parse_mode="HTML",
    )
    await state.set_state(AdminStates.waiting_report_start_date)
    await callback.answer()


@router.message(AdminStates.waiting_report_start_date)
async def lreport_get_start_date(message: Message, state: FSMContext):
    raw = message.text.strip()
    for fmt in ("%d.%m.%Y", "%Y-%m-%d"):
        try:
            d = datetime.strptime(raw, fmt).date()
            await state.update_data(report_start=d.isoformat())
            await message.answer(
                f"Начало: <b>{d.strftime('%d.%m.%Y')}</b>\n\nТеперь введите <b>дату конца</b> периода:",
                parse_mode="HTML",
            )
            await state.set_state(AdminStates.waiting_report_end_date)
            return
        except ValueError:
            pass
    await message.answer("Неверный формат. Попробуй ДД.ММ.ГГГГ, например: 01.05.2025")


@router.message(AdminStates.waiting_report_end_date)
async def lreport_get_end_date(message: Message, state: FSMContext):
    raw = message.text.strip()
    for fmt in ("%d.%m.%Y", "%Y-%m-%d"):
        try:
            d = datetime.strptime(raw, fmt).date()
            data = await state.get_data()
            await state.clear()
            start = data.get("report_start", d.isoformat())
            teachers = get_teachers_with_lessons()
            period_key = f"custom_{start}_{d.isoformat()}"
            await message.answer(
                f"Период: <b>{start} — {d.isoformat()}</b>\nВыберите преподавателя:",
                parse_mode="HTML",
                reply_markup=get_lessons_report_teacher_filter_keyboard(
                    teachers, period_key, back_callback="admin_teacher_lessons_report"
                ),
            )
            return
        except ValueError:
            pass
    await message.answer("Неверный формат. Попробуй ДД.ММ.ГГГГ, например: 31.05.2025")


@router.callback_query(lambda c: c.data.startswith("lreport_teacher_"))
async def lreport_show_report(callback: CallbackQuery):
    if not is_admin_role(callback.from_user.id):
        await callback.answer("Нет доступа", show_alert=True)
        return

    # parse callback: lreport_teacher_{all|ID}_{period_key}
    parts = callback.data.removeprefix("lreport_teacher_").split("_", 1)
    teacher_raw, period_key = parts[0], parts[1] if len(parts) > 1 else "week"
    teacher_id = None if teacher_raw == "all" else int(teacher_raw)

    if period_key.startswith("custom_"):
        _, start_date, end_date = period_key.split("_", 2)
        period_label = f"с {start_date} по {end_date}"
    else:
        start_date, end_date, period_label = _period_dates(period_key)

    rows = get_teacher_lessons_report(teacher_id=teacher_id, start_date=start_date, end_date=end_date)
    text = _format_report_text(rows, period_label, teacher_id)

    export_callback = f"lreport_excel_{teacher_raw}_{period_key}"
    export_kb = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="📥 Скачать Excel", callback_data=export_callback)],
        [InlineKeyboardButton(text="← Назад", callback_data="admin_teacher_lessons_report")],
    ])

    chunks = [text[i:i+3500] for i in range(0, len(text), 3500)]
    for i, chunk in enumerate(chunks):
        kb = export_kb if i == len(chunks) - 1 else None
        await callback.message.answer(chunk, parse_mode="HTML", reply_markup=kb)

    await callback.answer()


@router.callback_query(lambda c: c.data.startswith("lreport_excel_"))
async def lreport_export_excel(callback: CallbackQuery):
    if not is_admin_role(callback.from_user.id):
        await callback.answer("Нет доступа", show_alert=True)
        return

    parts = callback.data.removeprefix("lreport_excel_").split("_", 1)
    teacher_raw, period_key = parts[0], parts[1] if len(parts) > 1 else "week"
    teacher_id = None if teacher_raw == "all" else int(teacher_raw)

    if period_key.startswith("custom_"):
        _, start_date, end_date = period_key.split("_", 2)
        period_label = f"{start_date}_{end_date}"
    else:
        start_date, end_date, period_label = _period_dates(period_key)

    rows = get_teacher_lessons_report(teacher_id=teacher_id, start_date=start_date, end_date=end_date)
    if not rows:
        await callback.answer("Нет данных для выгрузки.", show_alert=True)
        return

    await callback.answer("Формирую файл...")
    try:
        xlsx_bytes = await asyncio.to_thread(_make_excel, rows, period_label)
        from aiogram.types import BufferedInputFile
        filename = f"lessons_{period_label.replace(' ', '_')}.xlsx"
        await callback.message.answer_document(
            BufferedInputFile(xlsx_bytes, filename=filename),
            caption=f"📊 Отчёт по занятиям {period_label}",
        )
    except Exception as exc:
        logger.error("Excel export error: %s", exc)
        await callback.message.answer("Ошибка при формировании Excel.")


@router.callback_query(lambda c: c.data == "sheets_refresh_all")
async def sheets_refresh_all(callback: CallbackQuery):
    if not is_admin_role(callback.from_user.id):
        await callback.answer("Нет доступа", show_alert=True)
        return
    client = get_sheets_client()
    if not client.is_configured():
        await callback.answer("Google Sheets не настроен.", show_alert=True)
        return
    await callback.answer("Обновляю таблицу, это займёт ~10 сек...")
    await callback.message.answer("🔄 Обновляю все листы таблицы...")
    asyncio.create_task(_do_sheets_refresh(callback.message))


async def _do_sheets_refresh(message) -> None:
    async with _summary_sheets_lock:
        results = []
        client = get_sheets_client()
        for fetch_fn, update_fn, label in [
            (get_weekly_payouts,       client.update_payouts_sheet,  "Выплаты"),
            (get_all_student_balances, client.update_balances_sheet, "Балансы"),
            (get_attendance_stats,     client.update_stats_sheet,    "Статистика"),
            (get_revenue_by_period,    client.update_revenue_sheet,  "Выручка"),
            (get_topups_history,       client.update_topups_sheet,   "Пополнения"),
        ]:
            try:
                data = await asyncio.to_thread(fetch_fn)
                ok = await asyncio.to_thread(update_fn, data)
                results.append(f"{'✅' if ok else '⚠️'} {label}")
            except Exception as exc:
                logger.warning("Sheets refresh %s failed: %s", label, exc)
                results.append(f"❌ {label}: ошибка")
        await message.answer(
            "📊 <b>Таблица обновлена:</b>\n" + "\n".join(results),
            parse_mode="HTML",
        )


_SHEET_TARGETS = {
    "sheets_refresh_balances": (get_all_student_balances, "update_balances_sheet", "Балансы"),
    "sheets_refresh_payouts":  (get_weekly_payouts,       "update_payouts_sheet",  "Выплаты"),
    "sheets_refresh_stats":    (get_attendance_stats,     "update_stats_sheet",    "Статистика"),
    "sheets_refresh_revenue":  (get_revenue_by_period,    "update_revenue_sheet",  "Выручка"),
    "sheets_refresh_topups":   (get_topups_history,       "update_topups_sheet",   "Пополнения"),
}


@router.callback_query(lambda c: c.data in _SHEET_TARGETS)
async def sheets_refresh_single(callback: CallbackQuery):
    if not is_admin_role(callback.from_user.id):
        await callback.answer("Нет доступа", show_alert=True)
        return
    client = get_sheets_client()
    if not client.is_configured():
        await callback.answer("Google Sheets не настроен.", show_alert=True)
        return
    fetch_fn, update_method, label = _SHEET_TARGETS[callback.data]
    await callback.answer(f"Обновляю {label}...")

    async def _do():
        async with _summary_sheets_lock:
            try:
                data = await asyncio.to_thread(fetch_fn)
                ok = await asyncio.to_thread(getattr(client, update_method), data)
                icon = "✅" if ok else "⚠️"
                await callback.message.answer(
                    f"{icon} <b>{label}</b> обновлён.", parse_mode="HTML"
                )
            except Exception as exc:
                logger.warning("Sheets single refresh %s failed: %s", label, exc)
                await callback.message.answer(f"❌ <b>{label}</b>: ошибка", parse_mode="HTML")

    asyncio.create_task(_do())


@router.message(Command("sheets_clear_dead"))
async def cmd_sheets_clear_dead(message: Message):
    if not is_admin_role(message.from_user.id):
        await message.answer("Нет доступа.")
        return
    dead = await asyncio.to_thread(sheets_outbox_get_dead)
    if not dead:
        await message.answer("✅ Застрявших записей нет.")
        return
    deleted = await asyncio.to_thread(sheets_outbox_delete_dead)
    await message.answer(
        f"🗑 Удалено <b>{deleted}</b> застрявших записей из outbox.\n"
        f"Данные за эти отметки можно восстановить вручную через бэкфил: "
        f"<code>python scripts/backfill_attendance_to_sheets.py --reformat</code>",
        parse_mode="HTML",
    )


@router.message(AdminStates.waiting_history_student_search)
async def show_balance_history(message: Message, state: FSMContext):
    if not is_admin_role(message.from_user.id):
        await message.answer("Нет доступа.")
        await state.clear()
        return

    search_text = message.text.strip()
    students = find_students_by_name_with_username(search_text)

    if not students:
        await message.answer("Ученик не найден.", reply_markup=get_admin_menu())
        await state.clear()
        return

    if len(students) > 1:
        await message.answer(
            "Найдено несколько учеников. Выберите нужного:",
            reply_markup=get_student_disambiguation_keyboard(students, action_prefix="history_pick_student"),
        )
        await state.clear()
        return

    student_id, full_name, telegram_id, phone, telegram_username = students[0]
    history_rows = get_balance_history_by_student(student_id)

    if not history_rows:
        await message.answer("История операций пока пустая.", reply_markup=get_admin_menu())
        await state.clear()
        return

    await send_student_contact_shortcut(
        message,
        telegram_id=telegram_id,
        username=telegram_username,
    )
    contact_keyboard = None
    if contact_keyboard:
        await message.answer("Быстрый переход в чат с учеником:", reply_markup=contact_keyboard)

    chunks = []
    current_chunk = [f"📘 <b>История баланса</b>\n\n👤 <b>{full_name}</b>\n"]

    for row in history_rows:
        _, student_name, teacher_name, subject_name, operation_type, lessons_delta, comment, created_at, created_by = row

        if operation_type == "manual_topup":
            op_text = "Начисление"
        elif operation_type == "attendance_writeoff":
            op_text = "Списание за посещение"
        else:
            op_text = operation_type

        sign = "+" if lessons_delta > 0 else ""

        entry = (
            f"\n📅 <b>{created_at}</b>\n"
            f"📚 {subject_name} — {teacher_name}\n"
            f"🧾 {op_text}\n"
            f"🔢 {sign}{lessons_delta}\n"
            f"💬 {comment if comment else '-'}\n"
            f"👨‍💼 ID кто сделал: {created_by if created_by else '-'}\n"
        )

        current_chunk.append(entry)

        if sum(len(x) for x in current_chunk) > 3000:
            chunks.append("".join(current_chunk))
            current_chunk = []

    if current_chunk:
        chunks.append("".join(current_chunk))

    for chunk in chunks:
        await message.answer(chunk, parse_mode="HTML")

    await message.answer("История показана.", reply_markup=get_admin_menu())
    await state.clear()


@router.callback_query(lambda c: c.data.startswith("history_pick_student_"))
async def history_pick_student_from_disambiguation(callback: CallbackQuery):
    if not is_admin_role(callback.from_user.id):
        await callback.answer("Нет доступа", show_alert=True)
        return

    try:
        student_id = int(callback.data.split("_")[-1])
    except (TypeError, ValueError):
        await callback.answer("Не удалось определить ученика", show_alert=True)
        return

    student = get_student_by_id_with_username(student_id)
    if not student:
        await callback.answer("Ученик не найден", show_alert=True)
        return

    _id, full_name, telegram_id, _phone, telegram_username = student
    history_rows = get_balance_history_by_student(student_id)
    if not history_rows:
        await callback.message.answer("История операций пока пустая.", reply_markup=get_admin_reply_menu(callback.from_user.id))
        await callback.answer()
        return

    await send_student_contact_shortcut(
        callback.message,
        telegram_id=telegram_id,
        username=telegram_username,
    )
    contact_keyboard = None
    if contact_keyboard:
        await callback.message.answer("Быстрый переход в чат с учеником:", reply_markup=contact_keyboard)

    chunks = []
    current_chunk = [f"📘 <b>История баланса</b>\n\n👤 <b>{full_name}</b>\n"]

    for row in history_rows:
        _, _student_name, teacher_name, subject_name, operation_type, lessons_delta, comment, created_at, created_by = row

        if operation_type == "manual_topup":
            op_text = "Начисление"
        elif operation_type == "attendance_writeoff":
            op_text = "Списание за посещение"
        else:
            op_text = operation_type

        sign = "+" if lessons_delta > 0 else ""

        entry = (
            f"\n📅 <b>{created_at}</b>\n"
            f"📚 {subject_name} — {teacher_name}\n"
            f"🧾 {op_text}\n"
            f"🔢 {sign}{lessons_delta}\n"
            f"💬 {comment if comment else '-'}\n"
            f"👨‍💼 ID кто сделал: {created_by if created_by else '-'}\n"
        )

        current_chunk.append(entry)

        if sum(len(x) for x in current_chunk) > 3000:
            chunks.append("".join(current_chunk))
            current_chunk = []

    if current_chunk:
        chunks.append("".join(current_chunk))

    for chunk in chunks:
        await callback.message.answer(chunk, parse_mode="HTML")

    await callback.message.answer("История показана.", reply_markup=get_admin_reply_menu(callback.from_user.id))
    await callback.answer()


@router.callback_query(lambda c: c.data == "admin_delete_user")
async def admin_delete_user(callback: CallbackQuery, state: FSMContext):
    if not is_admin_role(callback.from_user.id):
        await callback.answer("Нет доступа", show_alert=True)
        return

    await state.clear()
    if callback.from_user.id in SUPERADMINS:
        role_hint = "Доступно удаление ролей: администратор, преподаватель, ученик."
    else:
        role_hint = "Доступно удаление ролей: преподаватель, ученик."

    await callback.message.answer(
        "Введите ФИО или @username пользователя для удаления.\n"
        f"{role_hint}",
        reply_markup=get_main_menu_shortcut_keyboard(),
    )
    await state.set_state(AdminStates.waiting_delete_user_query)
    await callback.answer()


@router.callback_query(lambda c: c.data == "admin_debt_report")
async def admin_debt_report(callback: CallbackQuery):
    if not is_admin_role(callback.from_user.id):
        await callback.answer("Нет доступа", show_alert=True)
        return

    overdue_days_raw = os.getenv("SCHOOL_DEBT_OVERDUE_DAYS", "7").strip()
    try:
        overdue_days = max(1, int(overdue_days_raw))
    except ValueError:
        overdue_days = 7
    report_data = build_daily_debt_report(
        report_date=date.today().isoformat(),
        overdue_days=overdue_days,
    )
    text = format_debt_report_text(report_data, overdue_days)
    await callback.message.answer(
        text,
        parse_mode="HTML",
        reply_markup=InlineKeyboardMarkup(
            inline_keyboard=[
                [InlineKeyboardButton(text="Открыть список должников", callback_data="admin_debtors")],
                [InlineKeyboardButton(text="Главное меню", callback_data="menu_home")],
            ]
        ),
    )
    await callback.answer("Отчёт сформирован")


@router.callback_query(lambda c: c.data == "admin_debtors")
async def admin_debtors(callback: CallbackQuery):
    if not is_admin_role(callback.from_user.id):
        await callback.answer("Нет доступа", show_alert=True)
        return

    debtors = get_current_debtors_summary(limit=200)
    if not debtors:
        await callback.message.answer(
            "Сейчас нет активных должников.",
            reply_markup=get_main_menu_shortcut_keyboard(),
        )
        await callback.answer()
        return

    duplicate_names: dict[str, int] = {}
    for item in debtors:
        full_name = (item.get("full_name") or "").strip()
        duplicate_names[full_name] = duplicate_names.get(full_name, 0) + 1

    duplicate_lines = [
        f"• {name} — {count} записей"
        for name, count in sorted(duplicate_names.items())
        if name and count > 1
    ]

    lines = [
        "Выберите ученика-должника:",
        "",
        "В карточке показывается @username или ID, чтобы различать одноимённые записи.",
    ]
    if duplicate_lines:
        lines.extend(
            [
                "",
                "Найдены одинаковые имена:",
                *duplicate_lines[:10],
                "",
                "Рекомендуется уточнить ФИО (например, добавить фамилию/класс).",
            ]
        )

    buttons: list[list[InlineKeyboardButton]] = []
    for debtor in debtors[:40]:
        student_id = int(debtor["student_id"])
        full_name = str(debtor.get("full_name") or f"Ученик #{student_id}")
        username = debtor.get("telegram_username")
        total_debt = int(debtor.get("total_debt_lessons") or 0)
        suffix = f"@{username}" if username else f"ID:{student_id}"
        text = f"{full_name} | {suffix} | долг: {total_debt}"
        buttons.append(
            [InlineKeyboardButton(text=text[:64], callback_data=f"admin_debtor_{student_id}")]
        )
    buttons.append([InlineKeyboardButton(text="Найти ученика / переименовать", callback_data="admin_find_student")])
    buttons.append([InlineKeyboardButton(text="Главное меню", callback_data="menu_home")])

    await callback.message.answer(
        "\n".join(lines),
        reply_markup=InlineKeyboardMarkup(inline_keyboard=buttons),
    )
    await callback.answer()


@router.callback_query(lambda c: c.data.startswith("admin_debtor_legacy_"))
async def admin_debtor_details(callback: CallbackQuery):
    if not is_admin_role(callback.from_user.id):
        await callback.answer("Нет доступа", show_alert=True)
        return

    try:
        student_id = int(callback.data.split("_")[-1])
    except ValueError:
        await callback.answer("Некорректный выбор", show_alert=True)
        return

    details = get_debtor_student_details(student_id)
    if not details or not details.get("directions"):
        await callback.answer("У этого ученика сейчас нет активного долга", show_alert=True)
        return

    full_name = details.get("full_name") or f"Ученик #{student_id}"
    username = details.get("telegram_username")
    telegram_id = details.get("telegram_id")
    total_debt = int(details.get("total_debt_lessons") or 0)
    phone = details.get("phone") or "-"

    lines = [
        f"Должник: <b>{full_name}</b>",
        f"Username: @{username}" if username else "Username: не указан",
        f"Telegram ID: <code>{telegram_id}</code>" if telegram_id else "Telegram ID: не указан",
        f"Телефон: {phone}",
        f"Суммарный долг: <b>{total_debt} занятий</b>",
        "",
        "<b>Долг по направлениям:</b>",
    ]
    for row in details["directions"][:20]:
        lines.append(f"• {row['subject_name']} — {row['teacher_name']} | долг: {row['debt_lessons']}")

    detail_buttons: list[list[InlineKeyboardButton]] = []
    if telegram_id:
        detail_buttons.append([InlineKeyboardButton(text="Открыть чат в Telegram", url=f"tg://user?id={telegram_id}")])
    elif username:
        detail_buttons.append([InlineKeyboardButton(text="Открыть профиль", url=f"https://t.me/{username}")])
    detail_buttons.append([InlineKeyboardButton(text="← К списку должников", callback_data="admin_debtors")])
    detail_buttons.append([InlineKeyboardButton(text="Главное меню", callback_data="menu_home")])

    await callback.message.answer(
        "\n".join(lines),
        parse_mode="HTML",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=detail_buttons),
    )
    await callback.answer()


@router.callback_query(lambda c: c.data.startswith("admin_debtor_"))
async def admin_debtor_details_v2(callback: CallbackQuery):
    try:
        await callback.answer()

        if not is_admin_role(callback.from_user.id):
            await callback.answer("Нет доступа", show_alert=True)
            return

        try:
            student_id = int(callback.data.split("_")[-1])
        except (TypeError, ValueError):
            await callback.answer("Некорректный выбор", show_alert=True)
            return

        details = get_debtor_student_details(student_id)
        if not details or not details.get("directions"):
            await callback.answer("У этого ученика сейчас нет активного долга", show_alert=True)
            return

        full_name = details.get("full_name") or f"Ученик #{student_id}"
        username_raw = details.get("telegram_username")
        username = str(username_raw).strip().lstrip("@") if username_raw else None
        telegram_id = details.get("telegram_id")
        total_debt = int(details.get("total_debt_lessons") or 0)
        phone = details.get("phone") or "-"

        lines = [
            f"Должник: {full_name}",
            f"Username: @{username}" if username else "Username: не указан",
            f"Telegram ID: {telegram_id}" if telegram_id else "Telegram ID: не указан",
            f"Телефон: {phone}",
            f"Суммарный долг: {total_debt} занятий",
            "",
            "Долг по направлениям:",
        ]
        for row in details["directions"][:20]:
            lines.append(f"• {row['subject_name']} — {row['teacher_name']} | долг: {row['debt_lessons']}")

        detail_buttons: list[list[InlineKeyboardButton]] = []
        detail_buttons.append([InlineKeyboardButton(text="← К списку должников", callback_data="admin_debtors")])
        detail_buttons.append([InlineKeyboardButton(text="Главное меню", callback_data="menu_home")])

        await callback.message.answer(
            "\n".join(lines),
            reply_markup=InlineKeyboardMarkup(inline_keyboard=detail_buttons),
        )
    except Exception as exc:
        logger.exception("admin_debtor_details_v2 failed: %s", exc)
        try:
            await callback.answer("Ошибка открытия карточки должника", show_alert=True)
        except Exception:
            pass


@router.message(AdminStates.waiting_delete_user_query)
async def process_delete_user_query(message: Message, state: FSMContext):
    if not is_admin_role(message.from_user.id):
        await message.answer("Нет доступа.")
        await state.clear()
        return

    text = message.text.strip()
    if text.lower() in {"отмена", "cancel", "/menu"}:
        await state.clear()
        await message.answer("Удаление отменено.", reply_markup=get_admin_reply_menu(message.from_user.id))
        return

    allowed_roles = ("admin", "teacher", "student") if message.from_user.id in SUPERADMINS else ("teacher", "student")
    candidates = search_users_by_name_or_username(text, roles=allowed_roles, limit=20)
    if not candidates:
        await message.answer(
            "Ничего не найдено. Попробуйте другой запрос (ФИО или @username).",
            reply_markup=get_main_menu_shortcut_keyboard(),
        )
        return

    prepared = []
    for user_id, telegram_id, full_name, role, _is_active, telegram_username in candidates:
        if telegram_id in SUPERADMINS:
            continue
        prepared.append((user_id, full_name, role, telegram_username, str(telegram_id)))

    if not prepared:
        await message.answer("Подходящих пользователей для удаления не найдено.")
        return

    await state.set_state(AdminStates.waiting_delete_user_selection)
    await message.answer(
        "Выберите пользователя для удаления:",
        reply_markup=get_user_selection_keyboard(prepared, "delete_user_pick"),
    )


@router.callback_query(
    AdminStates.waiting_delete_user_selection,
    lambda c: c.data.startswith("delete_user_pick_"),
)
async def process_delete_user_selection(callback: CallbackQuery, state: FSMContext):
    if not is_admin_role(callback.from_user.id):
        await callback.answer("Нет доступа", show_alert=True)
        return

    try:
        user_id = int(callback.data.split("_")[-1])
    except ValueError:
        await callback.answer("Некорректный выбор", show_alert=True)
        return

    target_user = get_user_by_id(user_id)
    if not target_user:
        await callback.answer("Пользователь не найден", show_alert=True)
        return

    _, target_telegram_id, target_full_name, target_role, _is_active, target_username = target_user
    if target_telegram_id in SUPERADMINS:
        await callback.answer("Удалять супер-админа запрещено", show_alert=True)
        return
    if target_role not in {"admin", "teacher", "student"}:
        await callback.answer("Эта роль не поддерживается для удаления", show_alert=True)
        return
    if not can_delete_role(callback.from_user.id, target_role):
        await callback.answer("Недостаточно прав для удаления", show_alert=True)
        return

    before_snapshot = {
        "full_name": target_full_name,
        "role": target_role,
        "telegram_id": target_telegram_id,
        "telegram_username": target_username,
    }
    result = delete_user_with_related_data(target_role, target_telegram_id)
    if not result.get("ok"):
        await callback.answer("Не удалось удалить пользователя", show_alert=True)
        return

    log_admin_action(
        admin_telegram_id=callback.from_user.id,
        action_type="delete_user",
        target_type=target_role,
        target_id=target_telegram_id,
        details={
            "before": before_snapshot,
            "after": None,
            "result": result,
        },
        status="success",
    )

    await state.clear()
    await callback.message.answer(
        f"Пользователь удален.\n"
        f"Имя: {target_full_name}\n"
        f"Роль: {role_title(target_role)}\n"
        f"Username: @{target_username if target_username else '-'}",
        reply_markup=get_admin_reply_menu(callback.from_user.id),
    )
    await callback.answer("Готово")


@router.callback_query(lambda c: c.data == "superadmin_add_admin")
async def superadmin_add_admin(callback: CallbackQuery, state: FSMContext):
    if callback.from_user.id not in SUPERADMINS:
        await callback.answer("Нет доступа", show_alert=True)
        return

    await callback.message.answer(
        "Отправьте @username нового администратора (обязательно).\n"
        "Подсказка: нажмите «Главное меню», если хотите выйти из сценария.",
        reply_markup=get_main_menu_shortcut_keyboard(),
    )
    await state.set_state(AdminStates.waiting_new_admin_username)
    await callback.answer()


@router.callback_query(lambda c: c.data == "superadmin_change_role")
async def superadmin_change_role(callback: CallbackQuery, state: FSMContext):
    if callback.from_user.id not in SUPERADMINS:
        await callback.answer("Нет доступа", show_alert=True)
        return

    await state.clear()
    await callback.message.answer(
        "Введите ФИО или @username пользователя, у которого нужно изменить роль:",
        reply_markup=get_main_menu_shortcut_keyboard(),
    )
    await state.set_state(AdminStates.waiting_role_change_query)
    await callback.answer()


@router.message(AdminStates.waiting_role_change_query)
async def process_role_change_query(message: Message, state: FSMContext):
    if message.from_user.id not in SUPERADMINS:
        await message.answer("Нет доступа.")
        await state.clear()
        return

    text = message.text.strip()
    if text.lower() in {"отмена", "cancel", "/menu"}:
        await state.clear()
        await message.answer("Изменение роли отменено.", reply_markup=get_superadmin_menu())
        return

    candidates = search_users_by_name_or_username(text, roles=("admin", "teacher", "student"), limit=20)
    if not candidates:
        await message.answer(
            "Пользователи не найдены. Попробуйте другой запрос.",
            reply_markup=get_main_menu_shortcut_keyboard(),
        )
        return

    prepared = []
    for user_id, telegram_id, full_name, role, _is_active, username in candidates:
        if telegram_id in SUPERADMINS:
            continue
        prepared.append((user_id, full_name, role, username, str(telegram_id)))

    if not prepared:
        await message.answer("Подходящие пользователи для смены роли не найдены.")
        return

    await message.answer(
        "Выберите пользователя:",
        reply_markup=get_user_selection_keyboard(prepared, "role_user_pick"),
    )
    await state.set_state(AdminStates.waiting_role_change_selection)


@router.callback_query(
    AdminStates.waiting_role_change_selection,
    lambda c: c.data.startswith("role_user_pick_"),
)
async def process_role_change_user_pick(callback: CallbackQuery, state: FSMContext):
    if callback.from_user.id not in SUPERADMINS:
        await callback.answer("Нет доступа", show_alert=True)
        return

    try:
        user_id = int(callback.data.split("_")[-1])
    except ValueError:
        await callback.answer("Некорректный выбор", show_alert=True)
        return

    target_user = get_user_by_id(user_id)
    if not target_user:
        await callback.answer("Пользователь не найден", show_alert=True)
        return

    _, target_telegram_id, target_full_name, current_role, _is_active, target_username = target_user
    await state.update_data(
        role_change_target_id=target_telegram_id,
        role_change_target_full_name=target_full_name,
        role_change_target_current_role=current_role,
        role_change_target_username=target_username,
    )
    await callback.message.answer(
        "Выберите новую роль:",
        reply_markup=get_role_change_keyboard(),
    )
    await callback.answer()


@router.callback_query(
    AdminStates.waiting_role_change_selection,
    lambda c: c.data.startswith("role_set_"),
)
async def process_role_change_selection(callback: CallbackQuery, state: FSMContext):
    if callback.from_user.id not in SUPERADMINS:
        await callback.answer("Нет доступа", show_alert=True)
        return

    action = callback.data.replace("role_set_", "", 1)
    if action == "cancel":
        await state.clear()
        await callback.message.answer("Изменение роли отменено.", reply_markup=get_superadmin_menu())
        await callback.answer()
        return

    data = await state.get_data()
    target_telegram_id = data.get("role_change_target_id")
    if not target_telegram_id:
        await state.clear()
        await callback.answer("Целевой пользователь не найден", show_alert=True)
        return

    target_user = get_user_by_telegram_id(target_telegram_id)
    if not target_user:
        await state.clear()
        await callback.message.answer("Пользователь не найден.", reply_markup=get_superadmin_menu())
        await callback.answer()
        return

    target_full_name = data.get("role_change_target_full_name") or target_user[2]
    current_role = data.get("role_change_target_current_role") or target_user[3]

    if action == "disabled":
        changed = set_user_active(target_telegram_id, False)
        if not changed:
            await callback.answer("Не удалось отключить доступ", show_alert=True)
            return

        log_admin_action(
            admin_telegram_id=callback.from_user.id,
            action_type="change_role",
            target_type="user",
            target_id=target_telegram_id,
            details={
                "before": {"role": current_role, "is_active": True},
                "after": {"role": current_role, "is_active": False},
            },
            status="success",
        )
        await state.clear()
        await callback.message.answer(
            f"Доступ пользователя отключён.\nTelegram ID: {target_telegram_id}",
            reply_markup=get_superadmin_menu(),
        )
        await callback.answer("Готово")
        return

    if action not in {"admin", "teacher", "student"}:
        await callback.answer("Неизвестная роль", show_alert=True)
        return

    if action == "teacher":
        subjects = [item for item in get_teacher_catalog_subjects() if item]
        await state.update_data(
            role_change_target_id=target_telegram_id,
            role_change_target_full_name=target_full_name,
            role_change_target_current_role=current_role,
            role_teacher_subject_options=subjects,
        )
        if subjects:
            preview = ", ".join(subjects[:12])
            await callback.message.answer(
                "Настройка карточки преподавателя.\n\n"
                f"ФИО: {target_full_name}\n"
                f"Текущая роль: {current_role}\n\n"
                "Введите предмет (можно новый) или выберите из существующих:\n"
                f"{preview}"
            )
        else:
            await callback.message.answer(
                "Настройка карточки преподавателя.\n\n"
                f"ФИО: {target_full_name}\n"
                f"Текущая роль: {current_role}\n\n"
                "Введите предмет (например: Математика)."
            )
        await state.set_state(AdminStates.waiting_role_teacher_subject)
        await callback.answer()
        return

    changed = update_user_role(target_telegram_id, action)
    if not changed:
        await callback.answer("Не удалось изменить роль", show_alert=True)
        return

    log_admin_action(
        admin_telegram_id=callback.from_user.id,
        action_type="change_role",
        target_type="user",
        target_id=target_telegram_id,
        details={
            "before": {"role": current_role, "is_active": True},
            "after": {"role": action, "is_active": True},
        },
        status="success",
    )

    await state.clear()
    await callback.message.answer(
        f"Роль пользователя обновлена.\nTelegram ID: {target_telegram_id}\nНовая роль: {action}",
        reply_markup=get_superadmin_menu(),
    )
    await callback.answer("Готово")


@router.message(AdminStates.waiting_role_teacher_subject)
async def process_role_teacher_subject(message: Message, state: FSMContext):
    if message.from_user.id not in SUPERADMINS:
        await message.answer("Нет доступа.")
        await state.clear()
        return

    subject_name = (message.text or "").strip()
    if len(subject_name) < 2:
        await message.answer("Введите корректный предмет (минимум 2 символа).")
        return

    await state.update_data(role_teacher_subject=subject_name)
    await message.answer(
        "Введите описание преподавателя.\n"
        "Если описание пока не нужно — отправьте символ: -"
    )
    await state.set_state(AdminStates.waiting_role_teacher_description)


@router.message(AdminStates.waiting_role_teacher_description)
async def process_role_teacher_description(message: Message, state: FSMContext):
    if message.from_user.id not in SUPERADMINS:
        await message.answer("Нет доступа.")
        await state.clear()
        return

    text = (message.text or "").strip()
    description = None if text == "-" else text
    await state.update_data(role_teacher_description=description)
    await message.answer(
        "Отправьте фото преподавателя.\n"
        "Если фото пока не нужно — отправьте символ: -"
    )
    await state.set_state(AdminStates.waiting_role_teacher_photo)


@router.message(AdminStates.waiting_role_teacher_photo)
async def process_role_teacher_photo(message: Message, state: FSMContext):
    if message.from_user.id not in SUPERADMINS:
        await message.answer("Нет доступа.")
        await state.clear()
        return

    photo_path = None
    if message.photo:
        try:
            photo_path = await save_teacher_photo(message)
        except Exception as exc:
            logger.exception("Failed to save role-change teacher photo locally: %s", exc)
            await message.answer("Не удалось сохранить фото. Отправьте другое фото или '-' чтобы пропустить.")
            return
    else:
        text = (message.text or "").strip()
        if text != "-":
            await message.answer("Отправьте фото или '-' чтобы пропустить.")
            return

    data = await state.get_data()
    target_telegram_id = data.get("role_change_target_id")
    target_full_name = data.get("role_change_target_full_name")
    current_role = data.get("role_change_target_current_role")
    subject_name = data.get("role_teacher_subject")
    description = data.get("role_teacher_description")

    if not target_telegram_id or not target_full_name or not subject_name:
        await state.clear()
        await message.answer("Не удалось завершить смену роли. Попробуйте заново.", reply_markup=get_superadmin_menu())
        return

    changed = update_user_role(int(target_telegram_id), "teacher")
    if not changed:
        await state.clear()
        await message.answer("Не удалось изменить роль пользователя.", reply_markup=get_superadmin_menu())
        return

    teacher_id = add_or_update_teacher_profile(
        full_name=target_full_name,
        subject_name=subject_name,
        telegram_id=int(target_telegram_id),
        description=description,
        photo_path=photo_path,
    )

    log_admin_action(
        admin_telegram_id=message.from_user.id,
        action_type="change_role",
        target_type="user",
        target_id=int(target_telegram_id),
        details={
            "before": {"role": current_role, "is_active": True},
            "after": {
                "role": "teacher",
                "is_active": True,
                "teacher_id": teacher_id,
                "subject_name": subject_name,
                "description": description,
                "photo_path": photo_path,
            },
        },
        status="success",
    )

    await state.clear()
    await message.answer(
        "Роль пользователя обновлена и карточка преподавателя создана.\n\n"
        f"Telegram ID: {target_telegram_id}\n"
        f"ФИО: {target_full_name}\n"
        f"Предмет: {subject_name}\n"
        f"Описание: {'добавлено' if description else 'не указано'}\n"
        f"Фото: {'добавлено' if photo_path else 'не указано'}",
        reply_markup=get_superadmin_menu(),
    )


@router.message(AdminStates.waiting_new_admin_username)
async def process_new_admin_username(message: Message, state: FSMContext):
    if message.from_user.id not in SUPERADMINS:
        await message.answer("Нет доступа.")
        await state.clear()
        return

    text = message.text.strip()
    if not is_valid_username(text):
        await message.answer("Введите корректный @username в формате @example_user")
        return

    normalized_username = normalize_telegram_username(text)
    telegram_id = get_known_telegram_user_id_by_username(normalized_username)

    onboarding_text = ""
    target_id_for_log = None
    if telegram_id is not None:
        add_user(
            telegram_id=telegram_id,
            full_name=f"Admin @{normalized_username}",
            role="admin",
            telegram_username=normalized_username,
        )
        target_id_for_log = telegram_id
    else:
        token = create_onboarding_invite(
            role="admin",
            full_name=f"Admin @{normalized_username}",
            telegram_username=normalized_username,
            entity_type="user",
            entity_id=None,
            created_by=message.from_user.id,
        )
        link = build_onboarding_link(token)
        if link:
            onboarding_text = (
                "\n\nПользователь еще не писал школьному боту.\n"
                "Отправьте ему ссылку для автоматической выдачи роли admin:\n"
                f"{link}"
            )
        else:
            onboarding_text = (
                "\n\nПользователь еще не писал школьному боту, но ссылка не сформирована "
                "(проверьте SCHOOL_BOT_USERNAME в .env)."
            )

    log_admin_action(
        admin_telegram_id=message.from_user.id,
        action_type="add_admin",
        target_type="user",
        target_id=target_id_for_log,
        details={
            "before": None,
            "after": {
                "role": "admin",
                "telegram_username": normalized_username,
                "telegram_id": telegram_id,
            },
        },
        status="success",
    )

    await message.answer(
        "✅ Администратор добавлен.\n"
        f"Username: @{normalized_username}\n"
        f"Telegram ID: {telegram_id if telegram_id else 'будет определен автоматически'}"
        f"{onboarding_text}",
        reply_markup=get_superadmin_menu()
    )
    await state.clear()


@router.callback_query(lambda c: c.data == "superadmin_edit_teacher")
async def superadmin_edit_teacher(callback: CallbackQuery, state: FSMContext):
    if callback.from_user.id not in SUPERADMINS:
        await callback.answer("Нет доступа", show_alert=True)
        return

    await state.clear()
    await callback.message.answer(
        "Введите ФИО, предмет или @username преподавателя для редактирования карточки:",
        reply_markup=get_main_menu_shortcut_keyboard(),
    )
    await state.set_state(AdminStates.waiting_edit_teacher_query)
    await callback.answer()


@router.message(AdminStates.waiting_edit_teacher_query)
async def process_edit_teacher_query(message: Message, state: FSMContext):
    if message.from_user.id not in SUPERADMINS:
        await message.answer("Нет доступа.")
        await state.clear()
        return

    query = message.text.strip()
    if query.lower() in {"отмена", "cancel", "/menu"}:
        await state.clear()
        await message.answer("Редактирование преподавателя отменено.", reply_markup=get_superadmin_menu())
        return

    teachers = search_teacher_profiles(query, limit=20)
    if not teachers:
        await message.answer(
            "Преподаватели не найдены. Попробуйте другой запрос.",
            reply_markup=get_main_menu_shortcut_keyboard(),
        )
        return

    await message.answer(
        "Выберите преподавателя:",
        reply_markup=get_teacher_selection_keyboard(teachers),
    )
    await state.set_state(AdminStates.waiting_edit_teacher_selection)


@router.callback_query(
    AdminStates.waiting_edit_teacher_selection,
    lambda c: c.data.startswith("edit_teacher_pick_"),
)
async def process_edit_teacher_pick(callback: CallbackQuery, state: FSMContext):
    if callback.from_user.id not in SUPERADMINS:
        await callback.answer("Нет доступа", show_alert=True)
        return

    try:
        teacher_id = int(callback.data.split("_")[-1])
    except ValueError:
        await callback.answer("Некорректный выбор", show_alert=True)
        return

    teacher = get_teacher_profile_by_id(teacher_id)
    if not teacher:
        await callback.answer("Преподаватель не найден", show_alert=True)
        return

    (
        _teacher_id,
        telegram_id,
        full_name,
        subject_name,
        description,
        photo_path,
        telegram_username,
    ) = teacher

    await state.update_data(
        edit_teacher_id=teacher_id,
        edit_teacher_old={
            "full_name": full_name,
            "subject_name": subject_name,
            "description": description,
            "photo_path": photo_path,
            "telegram_id": telegram_id,
            "telegram_username": telegram_username,
        },
        edit_teacher_full_name=full_name,
        edit_teacher_subject=subject_name,
        edit_teacher_description=description,
        edit_teacher_photo=photo_path,
        edit_teacher_username=telegram_username,
        edit_teacher_telegram_id=telegram_id,
    )

    await callback.message.answer(
        "Текущая карточка преподавателя:\n"
        f"ФИО: {full_name}\n"
        f"Предмет: {subject_name if subject_name else '-'}\n"
        f"Описание: {'есть' if description else 'нет'}\n"
        f"Фото: {'есть' if photo_path else 'нет'}\n"
        f"Username: @{telegram_username if telegram_username else '-'}\n\n"
        "Введите новое ФИО или '-' чтобы оставить текущее.",
        reply_markup=get_main_menu_shortcut_keyboard(),
    )
    await state.set_state(AdminStates.waiting_edit_teacher_full_name)
    await callback.answer()


@router.message(AdminStates.waiting_edit_teacher_full_name)
async def process_edit_teacher_full_name(message: Message, state: FSMContext):
    if message.from_user.id not in SUPERADMINS:
        await message.answer("Нет доступа.")
        await state.clear()
        return

    text = message.text.strip()
    if text != "-":
        if len(text) < 3:
            await message.answer("Введите корректное ФИО или '-'.")
            return
        await state.update_data(edit_teacher_full_name=text)

    subjects = [item for item in get_teacher_catalog_subjects() if item]
    if subjects:
        await state.update_data(
            edit_teacher_subject_options=subjects,
            edit_teacher_subject_custom=False,
        )
        await message.answer(
            "Выберите основной предмет из списка или нажмите «Добавить новый предмет»:",
            reply_markup=get_edit_teacher_subject_picker_keyboard(subjects),
        )
    else:
        await state.update_data(
            edit_teacher_subject_options=[],
            edit_teacher_subject_custom=True,
        )
        await message.answer("Список предметов пуст. Введите новый основной предмет текстом.")
    await state.set_state(AdminStates.waiting_edit_teacher_subject)


@router.callback_query(
    AdminStates.waiting_edit_teacher_subject,
    lambda c: c.data.startswith("edit_teacher_subject_pick_") or c.data == "edit_teacher_subject_add_new",
)
async def process_edit_teacher_subject_pick(callback: CallbackQuery, state: FSMContext):
    if callback.from_user.id not in SUPERADMINS:
        await callback.answer("Нет доступа", show_alert=True)
        await state.clear()
        return

    data = await state.get_data()
    subjects = data.get("edit_teacher_subject_options") or []

    if callback.data == "edit_teacher_subject_add_new":
        await state.update_data(edit_teacher_subject_custom=True)
        await callback.message.answer("Введите новый основной предмет текстом:")
        await callback.answer()
        return

    try:
        subject_index = int(callback.data.split("_")[-1])
    except (TypeError, ValueError):
        await callback.answer("Не удалось определить предмет", show_alert=True)
        return

    if subject_index < 0 or subject_index >= len(subjects):
        await callback.answer("Предмет не найден в текущем списке", show_alert=True)
        return

    subject_name = subjects[subject_index]
    await state.update_data(edit_teacher_subject=subject_name, edit_teacher_subject_custom=False)
    await callback.message.answer(
        "Введите новое описание, '-' чтобы оставить текущее, или 'очистить' чтобы убрать описание."
    )
    await state.set_state(AdminStates.waiting_edit_teacher_description)
    await callback.answer()


@router.message(AdminStates.waiting_edit_teacher_subject)
async def process_edit_teacher_subject(message: Message, state: FSMContext):
    if message.from_user.id not in SUPERADMINS:
        await message.answer("Нет доступа.")
        await state.clear()
        return

    text = (message.text or "").strip()
    is_custom = bool((await state.get_data()).get("edit_teacher_subject_custom"))
    if not is_custom and text == "-":
        await message.answer("Выберите предмет кнопкой или нажмите «Добавить новый предмет».")
        return

    if text == "-":
        pass
    else:
        if len(text) < 2:
            await message.answer("Введите корректный предмет.")
            return
        await state.update_data(edit_teacher_subject=text, edit_teacher_subject_custom=True)

    await message.answer(
        "Введите новое описание, '-' чтобы оставить текущее, или 'очистить' чтобы убрать описание."
    )
    await state.set_state(AdminStates.waiting_edit_teacher_description)


@router.message(AdminStates.waiting_edit_teacher_description)
async def process_edit_teacher_description(message: Message, state: FSMContext):
    if message.from_user.id not in SUPERADMINS:
        await message.answer("Нет доступа.")
        await state.clear()
        return

    text = message.text.strip()
    if text.lower() == "очистить":
        await state.update_data(edit_teacher_description=None)
    elif text != "-":
        await state.update_data(edit_teacher_description=text)

    await message.answer(
        "Отправьте новое фото карточки.\n"
        "Отправьте '-' чтобы оставить текущее фото, или 'очистить' чтобы убрать фото."
    )
    await state.set_state(AdminStates.waiting_edit_teacher_photo)


@router.message(AdminStates.waiting_edit_teacher_photo)
async def process_edit_teacher_photo(message: Message, state: FSMContext):
    if message.from_user.id not in SUPERADMINS:
        await message.answer("Нет доступа.")
        await state.clear()
        return

    if message.photo:
        try:
            photo_path = await save_teacher_photo(message)
        except Exception as exc:
            logger.exception("Failed to save edited teacher photo locally: %s", exc)
            await message.answer("Не удалось сохранить фото. Попробуйте отправить изображение еще раз.")
            return
        await state.update_data(edit_teacher_photo=photo_path)
    else:
        text = message.text.strip()
        if text.lower() == "очистить":
            await state.update_data(edit_teacher_photo=None)
        elif text != "-":
            await message.answer("Отправьте фото, '-' или 'очистить'.")
            return

    await message.answer(
        "Введите новый @username преподавателя, '-' чтобы оставить текущий, или 'очистить' чтобы убрать."
    )
    await state.set_state(AdminStates.waiting_edit_teacher_username)


@router.message(AdminStates.waiting_edit_teacher_username)
async def process_edit_teacher_username(message: Message, state: FSMContext):
    if message.from_user.id not in SUPERADMINS:
        await message.answer("Нет доступа.")
        await state.clear()
        return

    data = await state.get_data()
    username_text = message.text.strip()

    current_username = data.get("edit_teacher_username")
    if username_text.lower() == "очистить":
        final_username = None
    elif username_text == "-":
        final_username = current_username
    else:
        if not is_valid_username(username_text):
            await message.answer("Введите корректный @username, '-' или 'очистить'.")
            return
        final_username = normalize_telegram_username(username_text)

    old = data.get("edit_teacher_old", {})
    teacher_id = data.get("edit_teacher_id")
    full_name = data.get("edit_teacher_full_name")
    subject_name = data.get("edit_teacher_subject")
    description = data.get("edit_teacher_description")
    photo_path = data.get("edit_teacher_photo")
    old_telegram_id = old.get("telegram_id")

    telegram_id = get_known_telegram_user_id_by_username(final_username)
    if telegram_id is None and final_username == old.get("telegram_username"):
        telegram_id = old_telegram_id

    updated = update_teacher_profile_fields(
        teacher_id,
        full_name=full_name,
        subject_name=subject_name,
        description=description,
        photo_path=photo_path,
    )
    if updated:
        set_teacher_telegram_id(teacher_id, telegram_id)

    onboarding_text = ""
    if final_username and telegram_id is None:
        token = create_onboarding_invite(
            role="teacher",
            full_name=full_name,
            telegram_username=final_username,
            entity_type="teacher",
            entity_id=teacher_id,
            created_by=message.from_user.id,
        )
        link = build_onboarding_link(token)
        if link:
            onboarding_text = (
                "\n\nПреподаватель еще не писал школьному боту.\n"
                "Отправьте ему ссылку для автоматической привязки:\n"
                f"{link}"
            )

    if telegram_id:
        add_user(
            telegram_id=telegram_id,
            full_name=full_name,
            role="teacher",
            telegram_username=final_username,
        )

    log_admin_action(
        admin_telegram_id=message.from_user.id,
        action_type="edit_teacher_profile",
        target_type="teacher",
        target_id=telegram_id,
        details={
            "before": old,
            "after": {
                "full_name": full_name,
                "subject_name": subject_name,
                "description": description,
                "photo_path": photo_path,
                "telegram_id": telegram_id,
                "telegram_username": final_username,
            },
        },
        status="success",
    )

    await state.clear()
    await message.answer(
        "Карточка преподавателя обновлена.\n"
        f"ФИО: {full_name}\n"
        f"Предмет: {subject_name}\n"
        f"Описание: {'есть' if description else 'нет'}\n"
        f"Фото: {'есть' if photo_path else 'нет'}\n"
        f"Username: @{final_username if final_username else '-'}\n"
        f"Telegram ID: {telegram_id if telegram_id else '-'}"
        f"{onboarding_text}",
        reply_markup=get_superadmin_menu(),
    )


@router.callback_query(lambda c: c.data == "superadmin_add_teacher")
async def superadmin_add_teacher(callback: CallbackQuery, state: FSMContext):
    if callback.from_user.id not in SUPERADMINS:
        await callback.answer("Нет доступа", show_alert=True)
        return

    await state.clear()
    await callback.message.answer(
        "Введите ФИО преподавателя.\n"
        "Подсказка: можно нажать «Главное меню» для выхода.",
        reply_markup=get_main_menu_shortcut_keyboard(),
    )
    await state.set_state(AdminStates.waiting_new_teacher_full_name)
    await callback.answer()


@router.message(AdminStates.waiting_new_teacher_full_name)
async def process_new_teacher_full_name(message: Message, state: FSMContext):
    if message.from_user.id not in SUPERADMINS:
        await message.answer("Нет доступа.")
        await state.clear()
        return

    full_name = (message.text or "").strip()
    if len(full_name) < 3:
        await message.answer("Введите корректное ФИО преподавателя.")
        return

    await state.update_data(new_teacher_full_name=full_name)
    subjects = [item for item in get_teacher_catalog_subjects() if item]
    if subjects:
        await state.update_data(new_teacher_subject_options=subjects, new_teacher_subject_custom=False)
        await message.answer(
            "Выберите предмет из списка или нажмите «Добавить новый предмет»:",
            reply_markup=get_teacher_subject_picker_keyboard(subjects),
        )
    else:
        await state.update_data(new_teacher_subject_options=[], new_teacher_subject_custom=True)
        await message.answer("Введите новый предмет для преподавателя:")
    await state.set_state(AdminStates.waiting_new_teacher_subject)


@router.callback_query(
    AdminStates.waiting_new_teacher_subject,
    lambda c: c.data.startswith("new_teacher_subject_pick_") or c.data == "new_teacher_subject_add_new",
)
async def process_new_teacher_subject_pick(callback: CallbackQuery, state: FSMContext):
    if callback.from_user.id not in SUPERADMINS:
        await callback.answer("Нет доступа", show_alert=True)
        await state.clear()
        return

    data = await state.get_data()
    subjects = data.get("new_teacher_subject_options") or []

    if callback.data == "new_teacher_subject_add_new":
        await state.update_data(new_teacher_subject_custom=True)
        await callback.message.answer("Введите новый предмет текстом:")
        await callback.answer()
        return

    try:
        subject_index = int(callback.data.split("_")[-1])
    except (TypeError, ValueError):
        await callback.answer("Не удалось определить предмет", show_alert=True)
        return

    if subject_index < 0 or subject_index >= len(subjects):
        await callback.answer("Предмет не найден", show_alert=True)
        return

    subject_name = subjects[subject_index]
    await state.update_data(new_teacher_subject=subject_name, new_teacher_subject_custom=False)
    await callback.message.answer("Введите описание преподавателя или отправьте '-' чтобы пропустить:")
    await state.set_state(AdminStates.waiting_new_teacher_description)
    await callback.answer()


@router.message(AdminStates.waiting_new_teacher_subject)
async def process_new_teacher_subject(message: Message, state: FSMContext):
    if message.from_user.id not in SUPERADMINS:
        await message.answer("Нет доступа.")
        await state.clear()
        return

    subject_name = (message.text or "").strip()
    if len(subject_name) < 2:
        await message.answer("Введите корректное название предмета.")
        return

    data = await state.get_data()
    subject_options = data.get("new_teacher_subject_options") or []
    normalized_lookup = {
        option.strip().lower(): option
        for option in subject_options
        if option and option.strip()
    }
    subject_name = normalized_lookup.get(subject_name.lower(), subject_name)

    await state.update_data(new_teacher_subject=subject_name)
    await message.answer("Введите описание преподавателя или отправьте '-' чтобы пропустить:")
    await state.set_state(AdminStates.waiting_new_teacher_description)


@router.message(AdminStates.waiting_new_teacher_description)
async def process_new_teacher_description(message: Message, state: FSMContext):
    if message.from_user.id not in SUPERADMINS:
        await message.answer("Нет доступа.")
        await state.clear()
        return

    text = (message.text or "").strip()
    description = None if text in {"-", "пропустить", "skip"} else text
    await state.update_data(new_teacher_description=description)
    await message.answer(
        "Отправьте фото карточки преподавателя или отправьте '-' чтобы пропустить.\n"
        "Можно использовать и новое фото, и текущую локальную картинку позже."
    )
    await state.set_state(AdminStates.waiting_new_teacher_photo)


@router.message(AdminStates.waiting_new_teacher_photo)
async def process_new_teacher_photo(message: Message, state: FSMContext):
    if message.from_user.id not in SUPERADMINS:
        await message.answer("Нет доступа.")
        await state.clear()
        return

    photo_path = None
    if message.photo:
        try:
            photo_path = await save_teacher_photo(message)
        except Exception as exc:
            logger.exception("Failed to save new teacher photo locally: %s", exc)
            await message.answer("Не удалось сохранить фото. Попробуйте отправить изображение еще раз.")
            return
    else:
        text = (message.text or "").strip()
        if text not in {"-", "пропустить", "skip"}:
            await message.answer("Отправьте фото или '-' для пропуска.")
            return

    await state.update_data(new_teacher_photo=photo_path)
    await message.answer("Теперь укажите @username преподавателя (обязательно):")
    await state.set_state(AdminStates.waiting_new_teacher_username)


@router.message(AdminStates.waiting_new_teacher_username)
async def process_new_teacher_username(message: Message, state: FSMContext):
    if message.from_user.id not in SUPERADMINS:
        await message.answer("Нет доступа.")
        await state.clear()
        return

    text = (message.text or "").strip()
    if not is_valid_username(text):
        await message.answer("Введите корректный @username в формате @example_user")
        return
    normalized_username = normalize_telegram_username(text)
    telegram_id = get_known_telegram_user_id_by_username(normalized_username)

    data = await state.get_data()
    teacher_name = data.get("new_teacher_full_name")
    subject_name = data.get("new_teacher_subject")
    description = data.get("new_teacher_description")
    photo_path = data.get("new_teacher_photo")

    if not teacher_name or not subject_name:
        await message.answer("Не удалось завершить создание преподавателя. Повторите снова.")
        await state.clear()
        return

    teacher_id = add_or_update_teacher_profile(
        full_name=teacher_name,
        subject_name=subject_name,
        telegram_id=telegram_id,
        description=description,
        photo_path=photo_path,
    )
    onboarding_text = ""
    if telegram_id:
        add_user(
            telegram_id=telegram_id,
            full_name=teacher_name,
            role="teacher",
            telegram_username=normalized_username,
        )
    else:
        token = create_onboarding_invite(
            role="teacher",
            full_name=teacher_name,
            telegram_username=normalized_username or "",
            entity_type="teacher",
            entity_id=teacher_id,
            created_by=message.from_user.id,
        )
        link = build_onboarding_link(token)
        if link:
            onboarding_text = (
                "\n\nПреподаватель еще не писал школьному боту.\n"
                "Отправьте ему ссылку для автоматической привязки роли teacher:\n"
                f"{link}"
            )
        else:
            onboarding_text = (
                "\n\nПреподаватель еще не писал школьному боту, но ссылка не сформирована "
                "(проверьте SCHOOL_BOT_USERNAME в .env)."
            )

    log_admin_action(
        admin_telegram_id=message.from_user.id,
        action_type="add_teacher",
        target_type="teacher",
        target_id=telegram_id,
        details={
            "before": None,
            "after": {
                "teacher_name": teacher_name,
                "subject_name": subject_name,
                "has_description": bool(description),
                "has_photo": bool(photo_path),
                "telegram_id": telegram_id,
                "telegram_username": normalized_username,
            },
        },
        status="success",
    )

    await message.answer(
        "✅ Преподаватель добавлен.\n\n"
        f"ФИО: {teacher_name}\n"
        f"Предмет: {subject_name}\n"
        f"Описание: {'добавлено' if description else 'пока нет'}\n"
        f"Фото карточки: {'добавлено' if photo_path else 'пока нет'}\n"
        f"Username: @{normalized_username}\n"
        f"Telegram ID: {telegram_id if telegram_id else 'будет определен автоматически'}"
        f"{onboarding_text}",
        reply_markup=get_superadmin_menu()
    )
    await state.clear()


@router.callback_query(lambda c: c.data == "superadmin_list_admins")
async def superadmin_list_admins(callback: CallbackQuery):
    if callback.from_user.id not in SUPERADMINS:
        await callback.answer("Нет доступа", show_alert=True)
        return

    admins = get_users_by_role("admin")

    if not admins:
        await callback.message.answer("Админов пока нет.")
        await callback.answer()
        return

    lines = ["<b>Список админов:</b>\n"]
    for user in admins:
        _, telegram_id, full_name, role, is_active = user
        lines.append(f"• {full_name} — <code>{telegram_id}</code>")

    await callback.message.answer("\n".join(lines), parse_mode="HTML")
    await callback.answer()


@router.callback_query(lambda c: c.data == "superadmin_list_teachers")
async def superadmin_list_teachers(callback: CallbackQuery):
    if callback.from_user.id not in SUPERADMINS:
        await callback.answer("Нет доступа", show_alert=True)
        return

    teachers = get_users_by_role("teacher")

    if not teachers:
        await callback.message.answer("Учителей пока нет.")
        await callback.answer()
        return

    lines = ["<b>Список учителей:</b>\n"]
    for user in teachers:
        _, telegram_id, full_name, role, is_active = user
        lines.append(f"• {full_name} — <code>{telegram_id}</code>")

    await callback.message.answer("\n".join(lines), parse_mode="HTML")
    await callback.answer()


@router.callback_query(lambda c: c.data == "student_profile")
async def student_profile(callback: CallbackQuery):
    user = get_user_by_telegram_id(callback.from_user.id)
    if not user:
        await callback.answer("Нет доступа", show_alert=True)
        return

    _, telegram_id, full_name, role, is_active = user

    if role != "student" or not is_active:
        await callback.answer("Нет доступа", show_alert=True)
        return

    student = get_student_by_telegram_id(callback.from_user.id)
    if not student:
        await callback.message.answer(
            "Профиль ученика пока не найден в базе.\n"
            "Обратись к администратору."
        )
        await callback.answer()
        return

    student_id, student_name, student_telegram_id, phone = student
    admin_contacts = get_active_admin_contacts()
    admin_links_text = ""
    if admin_contacts:
        admin_links = [
            (
                f"• <a href=\"https://t.me/{username}\">{full_name}</a>"
                if username
                else f"• <a href=\"tg://user?id={telegram_id}\">{full_name}</a>"
            )
            for telegram_id, full_name, username in admin_contacts
        ]
        admin_links_text = "\n\n<b>Напишите администратору:</b>\n" + "\n".join(admin_links)

    await callback.message.answer(
        f"👤 <b>Мой профиль</b>\n\n"
        f"📝 <b>Имя:</b> {student_name}\n"
        f"📱 <b>Телефон:</b> {phone if phone else '-'}\n"
        f"🆔 <b>Telegram ID:</b> <code>{student_telegram_id if student_telegram_id else '-'}</code>"
        f"{admin_links_text}",
        parse_mode="HTML"
    )
    await callback.answer()


@router.callback_query(lambda c: c.data == "student_payment_history")
async def student_payment_history(callback: CallbackQuery):
    user = get_user_by_telegram_id(callback.from_user.id)
    if not user:
        await callback.answer("Нет доступа", show_alert=True)
        return

    _, _, _, role, is_active = user

    if role != "student" or not is_active:
        await callback.answer("Нет доступа", show_alert=True)
        return

    student = get_student_by_telegram_id(callback.from_user.id)
    if not student:
        await callback.message.answer(
            "Профиль ученика пока не найден в базе.\n"
            "Пожалуйста, обратитесь к администратору."
        )
        await callback.answer()
        return

    _, student_name, _, _ = student
    payments = get_recent_payment_history_by_telegram_user(callback.from_user.id, limit=4)

    lines = [f"💳 <b>История оплат</b>\n\n👤 <b>{student_name}</b>\n"]

    if not payments:
        lines.append("\nИстория оплат пока отсутствует.")
    else:
        status_map = {
            "pending": "Ожидает проверки",
            "processing": "На проверке",
            "approved": "Подтверждена",
            "rejected": "Отклонена",
        }
        for index, payment in enumerate(payments, start=1):
            payment_id, status, caption_text, created_at, _updated_at, lessons_added = payment
            lines.append(
                f"\n{index}. Оплата #{payment_id}\n"
                f"Статус: <b>{status_map.get(status, status)}</b>\n"
                f"Дата: {created_at}\n"
                f"Начислено занятий: <b>{lessons_added}</b>\n"
                f"Комментарий: {caption_text if caption_text else '-'}"
            )

    await callback.message.answer("".join(lines), parse_mode="HTML")
    await callback.answer()


@router.callback_query(lambda c: c.data == "student_directions")
async def student_directions(callback: CallbackQuery):
    user = get_user_by_telegram_id(callback.from_user.id)
    if not user:
        await callback.answer("Нет доступа", show_alert=True)
        return

    _, telegram_id, full_name, role, is_active = user

    if role != "student" or not is_active:
        await callback.answer("Нет доступа", show_alert=True)
        return

    student = get_student_by_telegram_id(callback.from_user.id)
    if not student:
        await callback.message.answer(
            "Профиль ученика пока не найден в базе.\n"
            "Обратись к администратору."
        )
        await callback.answer()
        return

    student_id, student_name, student_telegram_id, phone = student
    directions = get_student_directions(student_id)

    if not directions:
        await callback.message.answer("У тебя пока нет активных направлений.")
        await callback.answer()
        return

    lines = [f"📚 <b>Мои направления</b>\n\n👤 <b>{student_name}</b>\n"]

    total_lessons = sum(direction[3] for direction in directions)
    lines.append(f"\n<b>Всего занятий на балансе:</b> {total_lessons}\n")

    for direction in directions:
        direction_id, teacher_name, subject_name, lesson_balance, tariff_type = direction
        tariff_text = "Разовое" if tariff_type == "single" else "Пакет"

        lines.append(
            f"\n<b>{subject_name}</b>\n"
            f"👨‍🏫 Преподаватель: {teacher_name}\n"
            f"🧾 Тариф: {tariff_text}\n"
            f"🔢 Остаток занятий: {lesson_balance}"
        )

    await callback.message.answer("".join(lines), parse_mode="HTML")
    await callback.answer()
    return

    query = (message.text or "").strip()
    if len(query) < 2:
        await message.answer("Введите минимум 2 символа для поиска преподавателя.")
        return

    teachers = search_teacher_profiles(query, limit=20)
    if not teachers:
        await message.answer(
            "Преподаватели не найдены. Попробуйте другой запрос (часть ФИО или @username)."
        )
        return

    await message.answer(
        "Выберите преподавателя из найденных:",
        reply_markup=get_teacher_selection_keyboard(teachers, action_prefix="assign_teacher_pick"),
    )
    await state.set_state(AdminStates.waiting_teacher_selection)
    return


# ── Promo codes ───────────────────────────────────────────────────────────────

def _promo_discount_str(p: dict) -> str:
    dv = p["discount_value"]
    val = int(dv) if float(dv) == int(float(dv)) else dv
    dtype = "%" if p["discount_type"] == "percent" else "₽"
    return f"{val}{dtype}"


def _promo_detail_text(p: dict) -> str:
    disc = _promo_discount_str(p)
    status = "✅ Активен" if p["active"] else "📦 В архиве"
    until = p["valid_until"][:10] if p.get("valid_until") else "бессрочно"
    atp = p.get("applies_to_packages", 0)
    scope = "Только разовые" if atp == 0 else ("Только пакеты" if atp == 1 else "Разовые + пакеты")
    uses_str = f"{p['used_count']}" + (f" из {p['max_uses']}" if p["max_uses"] else "")
    return (
        f"🎟 <b>{p['code']}</b>\n\n"
        f"Скидка: <b>{disc}</b>\n"
        f"Применяется: {scope}\n"
        f"Срок: <b>{until}</b>\n"
        f"Использован: <b>{uses_str}</b> раз\n"
        f"Статус: {status}"
    )


def _promo_detail_kb(p: dict, from_archive: bool = False) -> InlineKeyboardMarkup:
    pid = p["id"]
    back_cb = "admin_promo_archive" if from_archive else "admin_promo_list"
    back_label = "← К архиву" if from_archive else "← К списку"
    toggle_btn = (
        InlineKeyboardButton(text="✅ Активировать", callback_data=f"promo_activate_{pid}")
        if not p["active"] else
        InlineKeyboardButton(text="📦 В архив", callback_data=f"promo_archive_{pid}")
    )
    return InlineKeyboardMarkup(inline_keyboard=[
        [toggle_btn],
        [InlineKeyboardButton(text="🗑 Удалить", callback_data=f"promo_delete_confirm_{pid}")],
        [InlineKeyboardButton(text=back_label, callback_data=back_cb)],
    ])


def _active_promo_btn(p: dict) -> str:
    disc = _promo_discount_str(p)
    uses = f"{p['used_count']}" + (f"/{p['max_uses']}" if p["max_uses"] else "")
    atp2 = p.get("applies_to_packages", 0)
    pkg = " 📦" if atp2 == 1 else (" ✅" if atp2 == 2 else "")
    return f"✅  {p['code']} — {disc}{pkg}   ({uses} исп.)"


def _archive_promo_btn(p: dict) -> str:
    disc = _promo_discount_str(p)
    uses = f"{p['used_count']}" + (f"/{p['max_uses']}" if p["max_uses"] else "")
    return f"📦  {p['code']} — {disc}   ({uses} исп.)"


def _build_active_list_kb(promos: list[dict]) -> InlineKeyboardMarkup:
    active = [p for p in promos if p["active"]]
    archive_count = sum(1 for p in promos if not p["active"])
    rows = []
    for p in active:
        rows.append([InlineKeyboardButton(text=_active_promo_btn(p), callback_data=f"promo_detail_{p['id']}" )])
    rows.append([InlineKeyboardButton(text="➕ Создать промокод", callback_data="admin_promo_create")])
    if archive_count:
        rows.append([InlineKeyboardButton(text=f"📦 Архив ({archive_count})", callback_data="admin_promo_archive")])
    rows.append([InlineKeyboardButton(text="← Назад", callback_data="admin_section_finance")])
    return InlineKeyboardMarkup(inline_keyboard=rows)


def _build_archive_kb(promos: list[dict]) -> InlineKeyboardMarkup:
    archived = [p for p in promos if not p["active"]]
    rows = []
    for p in archived:
        rows.append([InlineKeyboardButton(text=_archive_promo_btn(p), callback_data=f"promo_detail_arch_{p['id']}" )])
    rows.append([InlineKeyboardButton(text="← К активным", callback_data="admin_promo_list")])
    return InlineKeyboardMarkup(inline_keyboard=rows)


@router.callback_query(lambda c: c.data == "admin_promo_list")
async def admin_promo_list(callback: CallbackQuery, state: FSMContext):
    await state.clear()
    promos = list_promo_codes()
    active = [p for p in promos if p["active"]]
    text = "🎟 <b>Промокоды</b>\n\n" + ("Выберите промокод для управления." if active else "Активных промокодов нет.")
    await callback.message.edit_text(text, parse_mode="HTML", reply_markup=_build_active_list_kb(promos))
    await callback.answer()


@router.callback_query(lambda c: c.data == "admin_promo_archive")
async def admin_promo_archive_list(callback: CallbackQuery):
    promos = list_promo_codes()
    archived = [p for p in promos if not p["active"]]
    text = f"📦 <b>Архив промокодов</b>\n\n" + (f"Промокодов: {len(archived)}" if archived else "Архив пуст.")
    await callback.message.edit_text(text, parse_mode="HTML", reply_markup=_build_archive_kb(promos))
    await callback.answer()


@router.callback_query(lambda c: c.data.startswith("promo_detail_arch_"))
async def promo_detail_from_archive(callback: CallbackQuery):
    try:
        pid = int(callback.data.split("promo_detail_arch_")[1])
    except (ValueError, IndexError):
        await callback.answer("Ошибка", show_alert=True)
        return
    p = get_promo_code_by_id(pid)
    if not p:
        await callback.answer("Промокод не найден", show_alert=True)
        return
    await callback.message.edit_text(_promo_detail_text(p), parse_mode="HTML", reply_markup=_promo_detail_kb(p, from_archive=True))
    await callback.answer()


@router.callback_query(lambda c: c.data.startswith("promo_detail_") and not c.data.startswith("promo_detail_arch_"))
async def promo_detail(callback: CallbackQuery):
    try:
        pid = int(callback.data.split("promo_detail_")[1])
    except (ValueError, IndexError):
        await callback.answer("Ошибка", show_alert=True)
        return
    p = get_promo_code_by_id(pid)
    if not p:
        await callback.answer("Промокод не найден", show_alert=True)
        return
    await callback.message.edit_text(_promo_detail_text(p), parse_mode="HTML", reply_markup=_promo_detail_kb(p))
    await callback.answer()


@router.callback_query(lambda c: c.data.startswith("promo_archive_"))
async def promo_archive(callback: CallbackQuery):
    pid = int(callback.data.split("promo_archive_")[1])
    deactivate_promo_code(pid)
    p = get_promo_code_by_id(pid)
    if not p:
        await callback.answer("Не найден", show_alert=True)
        return
    await callback.answer("Перемещён в архив")
    await callback.message.edit_text(_promo_detail_text(p), parse_mode="HTML", reply_markup=_promo_detail_kb(p, from_archive=True))


@router.callback_query(lambda c: c.data.startswith("promo_activate_"))
async def promo_activate(callback: CallbackQuery):
    pid = int(callback.data.split("promo_activate_")[1])
    activate_promo_code(pid)
    p = get_promo_code_by_id(pid)
    if not p:
        await callback.answer("Не найден", show_alert=True)
        return
    await callback.answer("Активирован ✅")
    await callback.message.edit_text(_promo_detail_text(p), parse_mode="HTML", reply_markup=_promo_detail_kb(p))


@router.callback_query(lambda c: c.data.startswith("promo_delete_confirm_"))
async def promo_delete_confirm(callback: CallbackQuery):
    pid = int(callback.data.split("promo_delete_confirm_")[1])
    p = get_promo_code_by_id(pid)
    if not p:
        await callback.answer("Не найден", show_alert=True)
        return
    kb = InlineKeyboardMarkup(inline_keyboard=[[
        InlineKeyboardButton(text="🗑 Да, удалить", callback_data=f"promo_delete_yes_{pid}"),
        InlineKeyboardButton(text="Отмена", callback_data=f"promo_detail_{pid}"),
    ]])
    await callback.message.edit_text(
        f"⚠️ Удалить промокод <b>{p['code']}</b>?\n\nЭто действие необратимо.",
        parse_mode="HTML", reply_markup=kb,
    )
    await callback.answer()


@router.callback_query(lambda c: c.data.startswith("promo_delete_yes_"))
async def promo_delete_yes(callback: CallbackQuery):
    pid = int(callback.data.split("promo_delete_yes_")[1])
    delete_promo_code(pid)
    await callback.answer("Удалён 🗑", show_alert=False)
    promos = list_promo_codes()
    active = [p for p in promos if p["active"]]
    text = "🎟 <b>Промокоды</b>\n\n" + ("Выберите промокод для управления." if active else "Активных промокодов нет.")
    await callback.message.edit_text(text, parse_mode="HTML", reply_markup=_build_active_list_kb(promos))


@router.callback_query(lambda c: c.data == "admin_promo_create")
async def admin_promo_create_start(callback: CallbackQuery, state: FSMContext):
    await state.clear()
    await callback.message.edit_text(
        "🎟 <b>Создание промокода — шаг 1 из 6</b>\n\n"
        "<b>Название промокода</b>\n"
        "Придумайте код, который ученик будет вводить в боте.\n\n"
        "Требования: только латиница и цифры, минимум 2 символа.\n"
        "Будет автоматически переведён в ВЕРХНИЙ регистр.\n\n"
        "Пример: <code>ЛЕТО25</code>, <code>СКИДКА10</code>, <code>VIP</code>",
        parse_mode="HTML",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[[
            InlineKeyboardButton(text="Отмена", callback_data="admin_promo_list")
        ]]),
    )
    await state.set_state(AdminStates.waiting_promo_code_text)
    await callback.answer()


@router.message(AdminStates.waiting_promo_code_text)
async def admin_promo_code_text(message: Message, state: FSMContext):
    code = (message.text or "").strip().upper()
    if not code or len(code) < 2:
        await message.answer("Код слишком короткий. Введите минимум 2 символа.")
        return
    await state.update_data(promo_code=code)
    await message.answer(
        f"✅ Код: <code>{code}</code>\n\n"
        f"🎟 <b>Создание промокода — шаг 2 из 6</b>\n\n"
        "<b>Тип скидки</b>\n\n"
        "• <b>Процент (%)</b> — скидка в процентах от суммы\n"
        "  Например: -20% от 1500₽ = ученик платит 1200₽\n\n"
        "• <b>Фиксированная сумма (₽)</b> — скидка фиксированной суммой\n"
        "  Например: -300₽ от 1500₽ = ученик платит 1200₽",
        parse_mode="HTML",
        reply_markup=get_promo_discount_type_kb(),
    )


@router.callback_query(lambda c: c.data in ("promo_type_percent", "promo_type_fixed_rub"))
async def admin_promo_type(callback: CallbackQuery, state: FSMContext):
    dtype = "percent" if callback.data == "promo_type_percent" else "fixed_rub"
    await state.update_data(promo_type=dtype)
    unit = "%" if dtype == "percent" else "₽"
    type_desc = (
        "Процент от суммы оплаты. Например, скидка 20% от 1500₽ = ученик платит 1200₽"
        if dtype == "percent" else
        "Фиксированная сумма в рублях. Например, скидка 300₽ от 1500₽ = ученик платит 1200₽"
    )
    await callback.message.edit_text(
        f"🎟 <b>Создание промокода — шаг 3 из 6</b>\n\n"
        f"<b>Размер скидки</b> ({type_desc})\n\n"
        f"Введите число без знака {unit}.\n"
        f"Пример: <code>20</code> (для 20{unit}) или <code>500</code> (для 500{unit})",
        parse_mode="HTML",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[[
            InlineKeyboardButton(text="Отмена", callback_data="admin_promo_list")
        ]]),
    )
    await state.set_state(AdminStates.waiting_promo_discount_value)
    await callback.answer()


@router.message(AdminStates.waiting_promo_discount_value)
async def admin_promo_discount_value(message: Message, state: FSMContext):
    try:
        val = float((message.text or "").strip().replace(",", "."))
        if val <= 0:
            raise ValueError
    except ValueError:
        await message.answer("Введите положительное число (например: 20 или 500).")
        return
    await state.update_data(promo_value=val)
    await message.answer(
        "🎟 <b>Создание промокода — шаг 4 из 6</b>\n\n"
        "<b>Срок действия</b>\n"
        "До какой даты ученики смогут активировать промокод?\n\n"
        "Введите дату в формате: <code>31.12.2026 23:59</code>\n"
        "Или нажмите «Бессрочный» — промокод будет работать без ограничения по времени.",
        parse_mode="HTML",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[[
            InlineKeyboardButton(text="Бессрочный", callback_data="promo_until_none")
        ]]),
    )
    await state.set_state(AdminStates.waiting_promo_valid_until)


@router.callback_query(lambda c: c.data == "promo_until_none")
async def admin_promo_until_none(callback: CallbackQuery, state: FSMContext):
    await state.update_data(promo_until=None)
    await callback.message.edit_text(
        "🎟 <b>Создание промокода — шаг 5 из 6</b>\n\n"
        "<b>Лимит активаций</b>\n"
        "Сколько раз этот промокод можно активировать (разными учениками)?\n\n"
        "Введите число. Например: <code>10</code> — первые 10 учеников смогут его использовать.\n\n"
        "Или нажмите «Без ограничений» — любое количество учеников.\n\n"
        "⚠️ Каждый ученик может использовать промокод только 1 раз.",
        parse_mode="HTML",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[[
            InlineKeyboardButton(text="Без ограничений", callback_data="promo_uses_none")
        ]]),
    )
    await state.set_state(AdminStates.waiting_promo_max_uses)
    await callback.answer()


@router.message(AdminStates.waiting_promo_valid_until)
async def admin_promo_valid_until(message: Message, state: FSMContext):
    text = (message.text or "").strip()
    if text.lower() in ("нет", "no", "-"):
        await state.update_data(promo_until=None)
        until_display = "бессрочно"
    else:
        parsed = None
        for fmt in ("%d.%m.%Y %H:%M", "%d.%m.%Y", "%Y-%m-%d %H:%M", "%Y-%m-%d"):
            try:
                dt = datetime.strptime(text, fmt)
                parsed = dt.strftime("%Y-%m-%d %H:%M:%S")
                break
            except ValueError:
                continue
        if not parsed:
            await message.answer(
                "Не удалось распознать дату. Используйте формат <code>31.12.2026 23:59</code>\n"
                "или введите <code>нет</code> для бессрочного.",
                parse_mode="HTML",
            )
            return
        await state.update_data(promo_until=parsed)
        until_display = parsed
    await message.answer(
        f"✅ Срок действия: {until_display}\n\n"
        "🎟 <b>Создание промокода — шаг 5 из 6</b>\n\n"
        "<b>Лимит активаций</b>\n"
        "Сколько раз этот промокод можно активировать (разными учениками)?\n\n"
        "Введите число. Например: <code>10</code> — первые 10 учеников смогут его использовать.\n\n"
        "Или нажмите «Без ограничений» — любое количество учеников.\n\n"
        "⚠️ Каждый ученик может использовать промокод только 1 раз.",
        parse_mode="HTML",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[[
            InlineKeyboardButton(text="Без ограничений", callback_data="promo_uses_none")
        ]]),
    )
    await state.set_state(AdminStates.waiting_promo_max_uses)


@router.callback_query(lambda c: c.data == "promo_uses_none")
async def admin_promo_uses_none(callback: CallbackQuery, state: FSMContext):
    await state.update_data(promo_max_uses=None)
    await _ask_promo_applies_to(callback.message, state, edit=True)
    await callback.answer()


@router.message(AdminStates.waiting_promo_max_uses)
async def admin_promo_max_uses(message: Message, state: FSMContext):
    text = (message.text or "").strip()
    if text.lower() in ("нет", "no", "-"):
        await state.update_data(promo_max_uses=None)
    else:
        try:
            n = int(text)
            if n <= 0:
                raise ValueError
            await state.update_data(promo_max_uses=n)
        except ValueError:
            await message.answer("Введите целое положительное число или «нет».")
            return
    await _ask_promo_applies_to(message, state, edit=False)


async def _ask_promo_applies_to(msg, state: FSMContext, edit: bool = False):
    text = (
        "🎟 <b>Создание промокода — шаг 6 из 6</b>\n\n"
        "<b>На что распространяется скидка?</b>\n\n"
        "• <b>Только разовые занятия</b> — скидка при оплате одного урока\n"
        "• <b>Только пакеты занятий</b> — скидка при покупке пакета (6, 12 уроков и т.д.)\n"
        "• <b>Разовые + пакеты</b> — скидка на любой вид оплаты"
    )
    kb = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="🔢 Только разовые занятия", callback_data="promo_applies_single")],
        [InlineKeyboardButton(text="📦 Только пакеты занятий", callback_data="promo_applies_packages")],
        [InlineKeyboardButton(text="✅ Разовые + пакеты", callback_data="promo_applies_all")],
    ])
    if edit:
        await msg.edit_text(text, parse_mode="HTML", reply_markup=kb)
    else:
        await msg.answer(text, parse_mode="HTML", reply_markup=kb)
    await state.set_state(AdminStates.waiting_promo_applies_to)


@router.callback_query(lambda c: c.data in ("promo_applies_single", "promo_applies_packages", "promo_applies_all"))
async def admin_promo_applies_to(callback: CallbackQuery, state: FSMContext):
    # applies_to_packages: 0=single only, 1=packages only, 2=all
    mapping = {"promo_applies_single": 0, "promo_applies_packages": 1, "promo_applies_all": 2}
    applies_to_packages = mapping.get(callback.data, 0)
    await state.update_data(promo_applies_to_packages=applies_to_packages)
    await _ask_promo_assign_student(callback.message, state, edit=True)
    await callback.answer()


async def _ask_promo_assign_student(msg, state: FSMContext, edit: bool = False):
    text = (
        "🎟 <b>Создание промокода — назначение ученику (необязательно)</b>\n\n"
        "Если промокод предназначен конкретному ученику — введите его имя, и он будет автоматически активирован.\n\n"
        "Если промокод общий (для всех, кто сам введёт код) — нажмите «Пропустить»."
    )
    kb = InlineKeyboardMarkup(inline_keyboard=[[
        InlineKeyboardButton(text="Пропустить — общий промокод", callback_data="promo_assign_skip")
    ]])
    if edit:
        await msg.edit_text(text, parse_mode="HTML", reply_markup=kb)
    else:
        await msg.answer(text, parse_mode="HTML", reply_markup=kb)
    await state.set_state(AdminStates.waiting_promo_assign_student)


@router.callback_query(lambda c: c.data == "promo_assign_skip")
async def admin_promo_assign_skip(callback: CallbackQuery, state: FSMContext):
    await _finalize_promo(callback.message, state, student_id=None, edit=True)
    await callback.answer()


@router.message(AdminStates.waiting_promo_assign_student)
async def admin_promo_assign_student(message: Message, state: FSMContext):
    query = (message.text or "").strip()
    students = find_students_by_name_with_username(query, limit=10)
    if not students:
        await message.answer("Ученик не найден. Попробуйте другое имя или нажмите «Пропустить».",
                             reply_markup=InlineKeyboardMarkup(inline_keyboard=[[
                                 InlineKeyboardButton(text="Пропустить", callback_data="promo_assign_skip")
                             ]]))
        return
    if len(students) == 1:
        await _finalize_promo(message, state, student_id=students[0][0], edit=False)
        return
    kb_rows = [[InlineKeyboardButton(text=s[1], callback_data=f"promo_pick_student_{s[0]}")] for s in students[:10]]
    kb_rows.append([InlineKeyboardButton(text="Пропустить", callback_data="promo_assign_skip")])
    await message.answer("Найдено несколько учеников, выберите:", reply_markup=InlineKeyboardMarkup(inline_keyboard=kb_rows))


@router.callback_query(lambda c: c.data.startswith("promo_pick_student_"))
async def admin_promo_pick_student(callback: CallbackQuery, state: FSMContext):
    student_id = int(callback.data.split("promo_pick_student_")[1])
    await _finalize_promo(callback.message, state, student_id=student_id, edit=True)
    await callback.answer()


async def _finalize_promo(msg, state: FSMContext, student_id: int | None, edit: bool):
    data = await state.get_data()
    code = data.get("promo_code", "")
    dtype = data.get("promo_type", "percent")
    value = float(data.get("promo_value", 0))
    until = data.get("promo_until")
    max_uses = data.get("promo_max_uses")
    applies_to_packages = data.get("promo_applies_to_packages", 0)
    await state.clear()

    promo_id = await asyncio.to_thread(create_promo_code, code, dtype, value, until, max_uses, applies_to_packages)
    if promo_id is None:
        text = f"❌ Промокод <code>{code}</code> уже существует."
        if edit:
            await msg.edit_text(text, parse_mode="HTML")
        else:
            await msg.answer(text, parse_mode="HTML")
        return

    assigned_name = ""
    if student_id is not None:
        student = await asyncio.to_thread(get_student_by_id, student_id)
        ok = await asyncio.to_thread(assign_promo_to_student, student_id, promo_id)
        if ok and student:
            assigned_name = f"\n👤 Назначен ученику: {student[1]}"

    unit = "%" if dtype == "percent" else "₽"
    val_str = int(value) if value == int(value) else value
    until_str = f"до {until}" if until else "бессрочно"
    uses_str = f"макс. {max_uses} использований" if max_uses else "без ограничений"

    text = (
        f"✅ <b>Промокод создан</b>\n\n"
        f"Код: <code>{code}</code>\n"
        f"Скидка: {val_str}{unit}\n"
        f"Действует: {until_str}\n"
        f"Использований: {uses_str}"
        f"{assigned_name}"
    )
    kb = InlineKeyboardMarkup(inline_keyboard=[[
        InlineKeyboardButton(text="← К списку промокодов", callback_data="admin_promo_list")
    ]])
    if edit:
        await msg.edit_text(text, parse_mode="HTML", reply_markup=kb)
    else:
        await msg.answer(text, parse_mode="HTML", reply_markup=kb)
